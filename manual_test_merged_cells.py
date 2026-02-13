import pandas as pd
from openpyxl import Workbook

# Create an Excel file with merged cells
wb = Workbook()
ws = wb.active
ws.merge_cells("A1:A3")
ws["A1"] = "Hello"
file_path = "test_merged.xlsx"
wb.save(file_path)

# Read Excel with your new feature
try:
    df = pd.read_excel(file_path, fill_merged_cells=True)
    print(df)
except TypeError:
    print("fill_merged_cells argument not recognized (make sure your code is updated).")
