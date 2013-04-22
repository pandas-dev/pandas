"""
Module contains tools for processing files into DataFrames or other objects
"""
from StringIO import StringIO
import re
from itertools import izip
from urlparse import urlparse
import csv

import numpy as np

from pandas.core.index import Index, MultiIndex
from pandas.core.frame import DataFrame
import datetime
import pandas.core.common as com
from pandas.util import py3compat
from pandas.io.date_converters import generic_parser

from pandas.util.decorators import Appender

import pandas.lib as lib
import pandas.tslib as tslib
import pandas._parser as _parser
from pandas.tseries.period import Period
import json


class DateConversionError(Exception):
    pass

_parser_params = """Also supports optionally iterating or breaking of the file
into chunks.

Parameters
----------
filepath_or_buffer : string or file handle / StringIO. The string could be
    a URL. Valid URL schemes include http, ftp, and file. For file URLs, a host
    is expected. For instance, a local file could be
    file ://localhost/path/to/table.csv
%s
lineterminator : string (length 1), default None
    Character to break file into lines. Only valid with C parser
quotechar : string
quoting : string
skipinitialspace : boolean, default False
    Skip spaces after delimiter
escapechar : string
dtype : Type name or dict of column -> type
    Data type for data or columns. E.g. {'a': np.float64, 'b': np.int32}
compression : {'gzip', 'bz2', None}, default None
    For on-the-fly decompression of on-disk data
dialect : string or csv.Dialect instance, default None
    If None defaults to Excel dialect. Ignored if sep longer than 1 char
    See csv.Dialect documentation for more details
header : int, default 0 if names parameter not specified, otherwise None
    Row to use for the column labels of the parsed DataFrame. Specify None if
    there is no header row.
skiprows : list-like or integer
    Row numbers to skip (0-indexed) or number of rows to skip (int)
    at the start of the file
index_col : int or sequence or False, default None
    Column to use as the row labels of the DataFrame. If a sequence is given, a
    MultiIndex is used. If you have a malformed file with delimiters at the end
    of each line, you might consider index_col=False to force pandas to _not_
    use the first column as the index (row names)
names : array-like
    List of column names to use. If file contains no header row, then you
    should explicitly pass header=None
prefix : string or None (default)
    Prefix to add to column numbers when no header, e.g 'X' for X0, X1, ...
na_values : list-like or dict, default None
    Additional strings to recognize as NA/NaN. If dict passed, specific
    per-column NA values
true_values : list
    Values to consider as True
false_values : list
    Values to consider as False
keep_default_na : bool, default True
    If na_values are specified and keep_default_na is False the default NaN
    values are overridden, otherwise they're appended to
parse_dates : boolean, list of ints or names, list of lists, or dict
    If True -> try parsing the index.
    If [1, 2, 3] -> try parsing columns 1, 2, 3 each as a separate date column.
    If [[1, 3]] -> combine columns 1 and 3 and parse as a single date column.
    {'foo' : [1, 3]} -> parse columns 1, 3 as date and call result 'foo'
keep_date_col : boolean, default False
    If True and parse_dates specifies combining multiple columns then
    keep the original columns.
date_parser : function
    Function to use for converting a sequence of string columns to an
    array of datetime instances. The default uses dateutil.parser.parser
    to do the conversion.
dayfirst : boolean, default False
    DD/MM format dates, international and European format
thousands : str, default None
    Thousands separator
comment : str, default None
    Indicates remainder of line should not be parsed
    Does not support line commenting (will return empty line)
decimal : str, default '.'
    Character to recognize as decimal point. E.g. use ',' for European data
nrows : int, default None
    Number of rows of file to read. Useful for reading pieces of large files
iterator : boolean, default False
    Return TextParser object
chunksize : int, default None
    Return TextParser object for iteration
skipfooter : int, default 0
    Number of line at bottom of file to skip
converters : dict. optional
    Dict of functions for converting values in certain columns. Keys can either
    be integers or column labels
verbose : boolean, default False
    Indicate number of NA values placed in non-numeric columns
delimiter : string, default None
    Alternative argument name for sep. Regular expressions are accepted.
encoding : string, default None
    Encoding to use for UTF when reading/writing (ex. 'utf-8')
squeeze : boolean, default False
    If the parsed data only contains one column then return a Series
na_filter: boolean, default True
    Detect missing value markers (empty strings and the value of na_values). In
    data without any NAs, passing na_filter=False can improve the performance
    of reading a large file

Returns
-------
result : DataFrame or TextParser
"""

_csv_sep = """sep : string, default ','
    Delimiter to use. If sep is None, will try to automatically determine
    this. Regular expressions are accepted.
"""

_table_sep = """sep : string, default \\t (tab-stop)
    Delimiter to use. Regular expressions are accepted."""

_read_csv_doc = """
Read CSV (comma-separated) file into DataFrame

%s
""" % (_parser_params % _csv_sep)

_read_table_doc = """
Read general delimited file into DataFrame

%s
""" % (_parser_params % _table_sep)

_fwf_widths = """\
colspecs : a list of pairs (tuples), giving the extents
    of the fixed-width fields of each line as half-open internals
    (i.e.,  [from, to[  ).
widths : a list of field widths, which can be used instead of
    'colspecs' if the intervals are contiguous.
"""

_read_fwf_doc = """
Read a table of fixed-width formatted lines into DataFrame

%s

Also, 'delimiter' is used to specify the filler character of the
fields if it is not spaces (e.g., '~').
""" % (_parser_params % _fwf_widths)


def _is_url(url):
    """
    Very naive check to see if url is an http(s), ftp, or file location.
    """
    parsed_url = urlparse(url)
    if parsed_url.scheme in ['http', 'file', 'ftp', 'https']:
        return True
    else:
        return False


def _read(filepath_or_buffer, kwds):
    "Generic reader of line files."
    encoding = kwds.get('encoding', None)
    skipfooter = kwds.pop('skipfooter', None)
    if skipfooter is not None:
        kwds['skip_footer'] = skipfooter

    if isinstance(filepath_or_buffer, basestring) and _is_url(filepath_or_buffer):
        from urllib2 import urlopen
        filepath_or_buffer = urlopen(filepath_or_buffer)
        if py3compat.PY3:  # pragma: no cover
            if encoding:
                errors = 'strict'
            else:
                errors = 'replace'
                encoding = 'utf-8'
            bytes = filepath_or_buffer.read()
            filepath_or_buffer = StringIO(bytes.decode(encoding, errors))

    if kwds.get('date_parser', None) is not None:
        if isinstance(kwds['parse_dates'], bool):
            kwds['parse_dates'] = True

    # Extract some of the arguments (pass chunksize on).
    iterator = kwds.pop('iterator', False)
    nrows = kwds.pop('nrows', None)
    chunksize = kwds.get('chunksize', None)

    # Create the parser.
    parser = TextFileReader(filepath_or_buffer, **kwds)

    if nrows is not None:
        return parser.read(nrows)
    elif chunksize or iterator:
        return parser

    return parser.read()

_parser_defaults = {
    'delimiter': None,

    'doublequote': True,
    'escapechar': None,
    'quotechar': '"',
    'quoting': csv.QUOTE_MINIMAL,
    'skipinitialspace': False,
    'lineterminator': None,

    'header': 'infer',
    'index_col': None,
    'names': None,
    'prefix': None,
    'skiprows': None,
    'na_values': None,
    'true_values': None,
    'false_values': None,
    'skip_footer': 0,
    'converters': None,

    'keep_default_na': True,
    'thousands': None,
    'comment': None,

    # 'engine': 'c',
    'parse_dates': False,
    'keep_date_col': False,
    'dayfirst': False,
    'date_parser': None,

    'usecols': None,

    # 'nrows': None,
    # 'iterator': False,
    'chunksize': None,
    'verbose': False,
    'encoding': None,
    'squeeze': False,
    'compression': None
}


_c_parser_defaults = {
    'delim_whitespace': False,
    'as_recarray': False,
    'na_filter': True,
    'compact_ints': False,
    'use_unsigned': False,
    'low_memory': True,
    'memory_map': False,
    'buffer_lines': None,
    'error_bad_lines': True,
    'warn_bad_lines': True,
    'factorize': True,
    'dtype': None,
    'decimal': b'.'
}

_fwf_defaults = {
    'colspecs': None,
    'widths': None
}

_c_unsupported = set(['skip_footer'])
_python_unsupported = set(_c_parser_defaults.keys())


def _make_parser_function(name, sep=','):

    def parser_f(filepath_or_buffer,
                 sep=sep,
                 dialect=None,
                 compression=None,

                 doublequote=True,
                 escapechar=None,
                 quotechar='"',
                 quoting=csv.QUOTE_MINIMAL,
                 skipinitialspace=False,
                 lineterminator=None,

                 header='infer',
                 index_col=None,
                 names=None,
                 prefix=None,
                 skiprows=None,
                 skipfooter=None,
                 skip_footer=0,
                 na_values=None,
                 true_values=None,
                 false_values=None,
                 delimiter=None,
                 converters=None,
                 dtype=None,
                 usecols=None,

                 engine='c',
                 delim_whitespace=False,
                 as_recarray=False,
                 na_filter=True,
                 compact_ints=False,
                 use_unsigned=False,
                 low_memory=_c_parser_defaults['low_memory'],
                 buffer_lines=None,
                 warn_bad_lines=True,
                 error_bad_lines=True,

                 keep_default_na=True,
                 thousands=None,
                 comment=None,
                 decimal=b'.',

                 parse_dates=False,
                 keep_date_col=False,
                 dayfirst=False,
                 date_parser=None,

                 memory_map=False,
                 nrows=None,
                 iterator=False,
                 chunksize=None,

                 verbose=False,
                 encoding=None,
                 squeeze=False):

        # Alias sep -> delimiter.
        if delimiter is None:
            delimiter = sep

        kwds = dict(delimiter=delimiter,
                    engine=engine,
                    dialect=dialect,
                    compression=compression,

                    doublequote=doublequote,
                    escapechar=escapechar,
                    quotechar=quotechar,
                    quoting=quoting,
                    skipinitialspace=skipinitialspace,
                    lineterminator=lineterminator,

                    header=header,
                    index_col=index_col,
                    names=names,
                    prefix=prefix,
                    skiprows=skiprows,
                    na_values=na_values,
                    true_values=true_values,
                    false_values=false_values,
                    keep_default_na=keep_default_na,
                    thousands=thousands,
                    comment=comment,
                    decimal=decimal,

                    parse_dates=parse_dates,
                    keep_date_col=keep_date_col,
                    dayfirst=dayfirst,
                    date_parser=date_parser,

                    nrows=nrows,
                    iterator=iterator,
                    chunksize=chunksize,
                    skipfooter=skipfooter or skip_footer,
                    converters=converters,
                    dtype=dtype,
                    usecols=usecols,
                    verbose=verbose,
                    encoding=encoding,
                    squeeze=squeeze,
                    memory_map=memory_map,

                    na_filter=na_filter,
                    compact_ints=compact_ints,
                    use_unsigned=use_unsigned,
                    delim_whitespace=delim_whitespace,
                    as_recarray=as_recarray,
                    warn_bad_lines=warn_bad_lines,
                    error_bad_lines=error_bad_lines,
                    low_memory=low_memory,
                    buffer_lines=buffer_lines)

        return _read(filepath_or_buffer, kwds)

    parser_f.__name__ = name

    return parser_f

read_csv = _make_parser_function('read_csv', sep=',')
read_csv = Appender(_read_csv_doc)(read_csv)

read_table = _make_parser_function('read_table', sep='\t')
read_table = Appender(_read_table_doc)(read_table)


@Appender(_read_fwf_doc)
def read_fwf(filepath_or_buffer, colspecs=None, widths=None, **kwds):
    # Check input arguments.
    if bool(colspecs is None) == bool(widths is None):
        raise ValueError("You must specify only one of 'widths' and "
                         "'colspecs'")

    # Compute 'colspec' from 'widths', if specified.
    if widths is not None:
        colspecs, col = [], 0
        for w in widths:
            colspecs.append((col, col + w))
            col += w

    kwds['colspecs'] = colspecs
    kwds['engine'] = 'python-fwf'
    return _read(filepath_or_buffer, kwds)


def read_clipboard(**kwargs):  # pragma: no cover
    """
    Read text from clipboard and pass to read_table. See read_table for the
    full argument list

    Returns
    -------
    parsed : DataFrame
    """
    from pandas.util.clipboard import clipboard_get
    text = clipboard_get()
    return read_table(StringIO(text), **kwargs)


def to_clipboard(obj):  # pragma: no cover
    """
    Attempt to write text representation of object to the system clipboard

    Notes
    -----
    Requirements for your platform
      - Linux: xsel command line tool
      - Windows: Python win32 extensions
      - OS X:
    """
    from pandas.util.clipboard import clipboard_set
    clipboard_set(str(obj))


# common NA values
# no longer excluding inf representations
# '1.#INF','-1.#INF', '1.#INF000000',
_NA_VALUES = set(['-1.#IND', '1.#QNAN', '1.#IND', '-1.#QNAN',
                 '#N/A N/A', 'NA', '#NA', 'NULL', 'NaN',
                 'nan', ''])


class TextFileReader(object):
    """

    Passed dialect overrides any of the related parser options

    """

    def __init__(self, f, engine='python', **kwds):

        self.f = f

        if kwds.get('dialect') is not None:
            dialect = kwds['dialect']
            kwds['delimiter'] = dialect.delimiter
            kwds['doublequote'] = dialect.doublequote
            kwds['escapechar'] = dialect.escapechar
            kwds['skipinitialspace'] = dialect.skipinitialspace
            kwds['quotechar'] = dialect.quotechar
            kwds['quoting'] = dialect.quoting

        if kwds.get('header', 'infer') == 'infer':
            kwds['header'] = 0 if kwds.get('names') is None else None

        self.orig_options = kwds

        # miscellanea
        self.engine = engine
        self._engine = None

        options = self._get_options_with_defaults(engine)

        self.chunksize = options.pop('chunksize', None)
        self.squeeze = options.pop('squeeze', False)

        # might mutate self.engine
        self.options, self.engine = self._clean_options(options, engine)
        if 'has_index_names' in kwds:
            self.options['has_index_names'] = kwds['has_index_names']

        self._make_engine(self.engine)

    def _get_options_with_defaults(self, engine):
        kwds = self.orig_options

        options = {}
        for argname, default in _parser_defaults.iteritems():
            if argname in kwds:
                value = kwds[argname]
            else:
                value = default

            options[argname] = value

        for argname, default in _c_parser_defaults.iteritems():
            if argname in kwds:
                value = kwds[argname]
                if engine != 'c' and value != default:
                    raise ValueError('%s is not supported with %s parser' %
                                     (argname, engine))
            options[argname] = value

        if engine == 'python-fwf':
            for argname, default in _fwf_defaults.iteritems():
                if argname in kwds:
                    value = kwds[argname]
                options[argname] = value

        return options

    def _clean_options(self, options, engine):
        result = options.copy()

        sep = options['delimiter']
        if (sep is None and not options['delim_whitespace']):
            if engine == 'c':
                print 'Using Python parser to sniff delimiter'
                engine = 'python'
        elif sep is not None and len(sep) > 1:
            # wait until regex engine integrated
            engine = 'python'

        # C engine not supported yet
        if engine == 'c':
            if options['skip_footer'] > 0:
                engine = 'python'

        if engine == 'c':
            for arg in _c_unsupported:
                del result[arg]

        if 'python' in engine:
            for arg in _python_unsupported:
                del result[arg]

        index_col = options['index_col']
        names = options['names']
        converters = options['converters']
        na_values = options['na_values']
        skiprows = options['skiprows']

        # really delete this one
        keep_default_na = result.pop('keep_default_na')

        if _is_index_col(index_col):
            if not isinstance(index_col, (list, tuple, np.ndarray)):
                index_col = [index_col]
        result['index_col'] = index_col

        names = list(names) if names is not None else names

        # type conversion-related
        if converters is not None:
            assert(isinstance(converters, dict))
        else:
            converters = {}

        # Converting values to NA
        na_values = _clean_na_values(na_values, keep_default_na)

        if com.is_integer(skiprows):
            skiprows = range(skiprows)
        skiprows = set() if skiprows is None else set(skiprows)

        # put stuff back
        result['index_col'] = index_col
        result['names'] = names
        result['converters'] = converters
        result['na_values'] = na_values
        result['skiprows'] = skiprows

        return result, engine

    def __iter__(self):
        try:
            while True:
                yield self.read(self.chunksize)
        except StopIteration:
            pass

    def _make_engine(self, engine='c'):
        if engine == 'c':
            self._engine = CParserWrapper(self.f, **self.options)
        else:
            if engine == 'python':
                klass = PythonParser
            elif engine == 'python-fwf':
                klass = FixedWidthFieldParser
            self._engine = klass(self.f, **self.options)

    def _failover_to_python(self):
        raise NotImplementedError

    def read(self, nrows=None):
        suppressed_warnings = False
        if nrows is not None:
            if self.options.get('skip_footer'):
                raise ValueError('skip_footer not supported for iteration')

            # # XXX hack
            # if isinstance(self._engine, CParserWrapper):
            #     suppressed_warnings = True
            #     self._engine.set_error_bad_lines(False)

        ret = self._engine.read(nrows)

        if self.options.get('as_recarray'):
            return ret

        index, columns, col_dict = ret

        # May alter columns / col_dict
        # index, columns, col_dict = self._create_index(col_dict, columns)

        df = DataFrame(col_dict, columns=columns, index=index)

        if self.squeeze and len(df.columns) == 1:
            return df[df.columns[0]]
        return df

    def _create_index(self, col_dict, columns):
        pass

    def get_chunk(self, size=None):
        if size is None:
            size = self.chunksize
        return self.read(nrows=size)

def _is_index_col(col):
    return col is not None and col is not False


class ParserBase(object):

    def __init__(self, kwds):
        self.names = kwds.get('names')
        self.orig_names = None
        self.prefix = kwds.pop('prefix', None)

        self.index_col = kwds.pop('index_col', None)
        self.index_names = None

        self.parse_dates = kwds.pop('parse_dates', False)
        self.date_parser = kwds.pop('date_parser', None)
        self.dayfirst = kwds.pop('dayfirst', False)
        self.keep_date_col = kwds.pop('keep_date_col', False)

        self.na_values = kwds.get('na_values')
        self.true_values = kwds.get('true_values')
        self.false_values = kwds.get('false_values')

        self._date_conv = _make_date_converter(date_parser=self.date_parser,
                                               dayfirst=self.dayfirst)

        self._name_processed = False

    @property
    def _has_complex_date_col(self):
        return (isinstance(self.parse_dates, dict) or
                (isinstance(self.parse_dates, list) and
                 len(self.parse_dates) > 0 and
                 isinstance(self.parse_dates[0], list)))

    def _should_parse_dates(self, i):
        if isinstance(self.parse_dates, bool):
            return self.parse_dates
        else:
            name = self.index_names[i]
            j = self.index_col[i]

            if np.isscalar(self.parse_dates):
                return (j == self.parse_dates) or (name == self.parse_dates)
            else:
                return (j in self.parse_dates) or (name in self.parse_dates)

    def _make_index(self, data, alldata, columns):
        if not _is_index_col(self.index_col) or len(self.index_col) == 0:
            index = None

        elif not self._has_complex_date_col:
            index = self._get_simple_index(alldata, columns)
            index = self._agg_index(index)

        elif self._has_complex_date_col:
            if not self._name_processed:
                (self.index_names, _,
                 self.index_col) = _clean_index_names(list(columns),
                                                      self.index_col)
                self._name_processed = True
            index = self._get_complex_date_index(data, columns)
            index = self._agg_index(index, try_parse_dates=False)

        return index

    _implicit_index = False

    def _get_simple_index(self, data, columns):
        def ix(col):
            if not isinstance(col, basestring):
                return col
            raise ValueError('Index %s invalid' % col)
        index = None

        to_remove = []
        index = []
        for idx in self.index_col:
            i = ix(idx)
            to_remove.append(i)
            index.append(data[i])

        # remove index items from content and columns, don't pop in
        # loop
        for i in reversed(sorted(to_remove)):
            data.pop(i)
            if not self._implicit_index:
                columns.pop(i)

        return index

    def _get_complex_date_index(self, data, col_names):
        def _get_name(icol):
            if isinstance(icol, basestring):
                return icol

            if col_names is None:
                raise ValueError(('Must supply column order to use %s as '
                                  'index') % str(icol))

            for i, c in enumerate(col_names):
                if i == icol:
                    return c

        index = None

        to_remove = []
        index = []
        for idx in self.index_col:
            name = _get_name(idx)
            to_remove.append(name)
            index.append(data[name])

        # remove index items from content and columns, don't pop in
        # loop
        for c in reversed(sorted(to_remove)):
            data.pop(c)
            col_names.remove(c)

        return index

    def _agg_index(self, index, try_parse_dates=True):
        arrays = []
        for i, arr in enumerate(index):

            if (try_parse_dates and self._should_parse_dates(i)):
                arr = self._date_conv(arr)

            col_na_values = self.na_values

            if isinstance(self.na_values, dict):
                col_name = self.index_names[i]
                if col_name is not None:
                    col_na_values = _get_na_values(col_name,
                                                   self.na_values)

            arr, _ = self._convert_types(arr, col_na_values)
            arrays.append(arr)

        index = MultiIndex.from_arrays(arrays, names=self.index_names)

        return index

    def _convert_to_ndarrays(self, dct, na_values, verbose=False,
                             converters=None):
        result = {}
        for c, values in dct.iteritems():
            conv_f = None if converters is None else converters.get(c, None)
            col_na_values = _get_na_values(c, na_values)
            coerce_type = True
            if conv_f is not None:
                values = lib.map_infer(values, conv_f)
                coerce_type = False
            cvals, na_count = self._convert_types(values, col_na_values,
                                                  coerce_type)
            result[c] = cvals
            if verbose and na_count:
                print 'Filled %d NA values in column %s' % (na_count, str(c))
        return result

    def _convert_types(self, values, na_values, try_num_bool=True):
        na_count = 0
        if issubclass(values.dtype.type, (np.number, np.bool_)):
            mask = lib.ismember(values, na_values)
            na_count = mask.sum()
            if na_count > 0:
                if com.is_integer_dtype(values):
                    values = values.astype(np.float64)
                np.putmask(values, mask, np.nan)
            return values, na_count

        if try_num_bool:
            try:
                result = lib.maybe_convert_numeric(values, na_values, False)
            except Exception:
                result = values
                if values.dtype == np.object_:
                    na_count = lib.sanitize_objects(result, na_values, False)
        else:
            result = values
            if values.dtype == np.object_:
                na_count = lib.sanitize_objects(values, na_values, False)

        if result.dtype == np.object_ and try_num_bool:
            result = lib.maybe_convert_bool(values,
                                            true_values=self.true_values,
                                            false_values=self.false_values)

        return result, na_count

    def _do_date_conversions(self, names, data):
        # returns data, columns
        if self.parse_dates is not None:
            data, names = _process_date_conversion(
                data, self._date_conv, self.parse_dates, self.index_col,
                self.index_names, names, keep_date_col=self.keep_date_col)

        return names, data

    def _exclude_implicit_index(self, alldata):

        if self._implicit_index:
            excl_indices = self.index_col

            data = {}
            offset = 0
            for i, col in enumerate(self.orig_names):
                while i + offset in excl_indices:
                    offset += 1
                data[col] = alldata[i + offset]
        else:
            data = dict((k, v) for k, v in izip(self.orig_names, alldata))

        return data


class CParserWrapper(ParserBase):
    """

    """

    def __init__(self, src, **kwds):
        self.kwds = kwds
        kwds = kwds.copy()

        self.as_recarray = kwds.get('as_recarray', False)
        ParserBase.__init__(self, kwds)

        if 'utf-16' in (kwds.get('encoding') or ''):
            if isinstance(src, basestring):
                src = open(src, 'rb')
            src = com.UTF8Recoder(src, kwds['encoding'])
            kwds['encoding'] = 'utf-8'

        # #2442
        kwds['allow_leading_cols'] = self.index_col is not False
        self._reader = _parser.TextReader(src, **kwds)

        # XXX
        self.usecols = self._reader.usecols

        passed_names = self.names is None

        if self._reader.header is None:
            self.names = None
        else:
            self.names = list(self._reader.header)

        if self.names is None:
            if self.prefix:
                self.names = ['X%d' % i
                              for i in range(self._reader.table_width)]
            else:
                self.names = range(self._reader.table_width)

        # XXX
        self._set_noconvert_columns()

        self.orig_names = self.names

        if not self._has_complex_date_col:
            if (self._reader.leading_cols == 0 and
                    _is_index_col(self.index_col)):

                self._name_processed = True
                (self.index_names, self.names,
                 self.index_col) = _clean_index_names(self.names,
                                                      self.index_col)

            if self._reader.header is None and not passed_names:
                self.index_names = [None] * len(self.index_names)

        self._implicit_index = self._reader.leading_cols > 0

    def _set_noconvert_columns(self):
        names = self.names

        def _set(x):
            if com.is_integer(x):
                self._reader.set_noconvert(x)
            else:
                self._reader.set_noconvert(names.index(x))

        if isinstance(self.parse_dates, list):
            for val in self.parse_dates:
                if isinstance(val, list):
                    for k in val:
                        _set(k)
                else:
                    _set(val)

    def set_error_bad_lines(self, status):
        self._reader.set_error_bad_lines(int(status))

    def read(self, nrows=None):
        if self.as_recarray:
            # what to do if there are leading columns?
            return self._reader.read(nrows)

        try:
            data = self._reader.read(nrows)
        except StopIteration:
            if nrows is None:
                return None, self.names, {}
            else:
                raise

        names = self.names

        if self._reader.leading_cols:
            if self._has_complex_date_col:
                raise NotImplementedError('file structure not yet supported')

            # implicit index, no index names
            arrays = []

            for i in range(self._reader.leading_cols):
                if self.index_col is None:
                    values = data.pop(i)
                else:
                    values = data.pop(self.index_col[i])

                values = self._maybe_parse_dates(values, i,
                                                 try_parse_dates=True)
                arrays.append(values)

            index = MultiIndex.from_arrays(arrays)

            if self.usecols is not None:
                names = self._filter_usecols(names)

            # rename dict keys
            data = sorted(data.items())
            data = dict((k, v) for k, (i, v) in zip(names, data))

            names, data = self._do_date_conversions(names, data)

        else:
            # rename dict keys
            data = sorted(data.items())

            # ugh, mutation
            names = list(self.orig_names)

            if self.usecols is not None:
                names = self._filter_usecols(names)

            # columns as list
            alldata = [x[1] for x in data]

            data = dict((k, v) for k, (i, v) in zip(names, data))

            names, data = self._do_date_conversions(names, data)
            index = self._make_index(data, alldata, names)

        return index, names, data

    def _filter_usecols(self, names):
        # hackish
        if self.usecols is not None and len(names) != len(self.usecols):
            names = [name for i, name in enumerate(names)
                     if i in self.usecols or name in self.usecols]
        return names

    def _get_index_names(self):
        names = list(self._reader.header)
        idx_names = None

        if self._reader.leading_cols == 0 and self.index_col is not None:
            (idx_names, names,
             self.index_col) = _clean_index_names(names, self.index_col)

        return names, idx_names

    def _maybe_parse_dates(self, values, index, try_parse_dates=True):
        if try_parse_dates and self._should_parse_dates(index):
            values = self._date_conv(values)
        return values


def TextParser(*args, **kwds):
    """
    Converts lists of lists/tuples into DataFrames with proper type inference
    and optional (e.g. string to datetime) conversion. Also enables iterating
    lazily over chunks of large files

    Parameters
    ----------
    data : file-like object or list
    delimiter : separator character to use
    dialect : str or csv.Dialect instance, default None
        Ignored if delimiter is longer than 1 character
    names : sequence, default
    header : int, default 0
        Row to use to parse column labels. Defaults to the first row. Prior
        rows will be discarded
    index_col : int or list, default None
        Column or columns to use as the (possibly hierarchical) index
    has_index_names: boolean, default False
        True if the cols defined in index_col have an index name and are
        not in the header
    na_values : iterable, default None
        Custom NA values
    keep_default_na : bool, default True
    thousands : str, default None
        Thousands separator
    comment : str, default None
        Comment out remainder of line
    parse_dates : boolean, default False
    keep_date_col : boolean, default False
    date_parser : function, default None
    skiprows : list of integers
        Row numbers to skip
    skip_footer : int
        Number of line at bottom of file to skip
    encoding : string, default None
        Encoding to use for UTF when reading/writing (ex. 'utf-8')
    squeeze : boolean, default False
        returns Series if only one column
    """
    kwds['engine'] = 'python'
    return TextFileReader(*args, **kwds)

# delimiter=None, dialect=None, names=None, header=0,
# index_col=None,
# na_values=None,
# na_filter=True,
# thousands=None,
# quotechar='"',
# escapechar=None,
# doublequote=True,
# skipinitialspace=False,
# quoting=csv.QUOTE_MINIMAL,
# comment=None, parse_dates=False, keep_date_col=False,
# date_parser=None, dayfirst=False,
# chunksize=None, skiprows=None, skip_footer=0, converters=None,
# verbose=False, encoding=None, squeeze=False):


def count_empty_vals(vals):
    return sum([1 for v in vals if v == '' or v is None])


def _wrap_compressed(f, compression):
    compression = compression.lower()
    if compression == 'gzip':
        import gzip
        return gzip.GzipFile(fileobj=f)
    elif compression == 'bz2':
        raise ValueError('Python cannot read bz2 data from file handle')
    else:
        raise ValueError('do not recognize compression method %s'
                         % compression)


class PythonParser(ParserBase):

    def __init__(self, f, **kwds):
        """
        Workhorse function for processing nested list into DataFrame

        Should be replaced by np.genfromtxt eventually?
        """
        ParserBase.__init__(self, kwds)

        self.data = None
        self.buf = []
        self.pos = 0

        if kwds['usecols'] is not None:
            raise Exception("usecols not supported with engine='python'"
                            " or multicharacter separators (yet).")

        self.header = kwds['header']
        self.encoding = kwds['encoding']
        self.compression = kwds['compression']
        self.skiprows = kwds['skiprows']

        self.skip_footer = kwds['skip_footer']
        self.delimiter = kwds['delimiter']

        self.quotechar = kwds['quotechar']
        self.escapechar = kwds['escapechar']
        self.doublequote = kwds['doublequote']
        self.skipinitialspace = kwds['skipinitialspace']
        self.lineterminator = kwds['lineterminator']
        self.quoting = kwds['quoting']

        self.has_index_names = False
        if 'has_index_names' in kwds:
            self.has_index_names = kwds['has_index_names']

        self.verbose = kwds['verbose']
        self.converters = kwds['converters']

        self.thousands = kwds['thousands']
        self.comment = kwds['comment']
        self._comment_lines = []

        if isinstance(f, basestring):
            f = com._get_handle(f, 'r', encoding=self.encoding,
                                compression=self.compression)
        elif self.compression:
            f = _wrap_compressed(f, self.compression)

        if hasattr(f, 'readline'):
            self._make_reader(f)
        else:
            self.data = f
        self.columns = self._infer_columns()

        # get popped off for index
        self.orig_names = list(self.columns)

        # needs to be cleaned/refactored
        # multiple date column thing turning into a real spaghetti factory

        if not self._has_complex_date_col:
            (self.index_names,
             self.orig_names, _) = self._get_index_name(self.columns)
            self._name_processed = True
        self._first_chunk = True

    def _make_reader(self, f):
        sep = self.delimiter

        if sep is None or len(sep) == 1:
            if self.lineterminator:
                raise ValueError('Custom line terminators not supported in '
                                 'python parser (yet)')

            class MyDialect(csv.Dialect):
                delimiter = self.delimiter
                quotechar = self.quotechar
                escapechar = self.escapechar
                doublequote = self.doublequote
                skipinitialspace = self.skipinitialspace
                quoting = self.quoting
                lineterminator = '\n'

            dia = MyDialect

            sniff_sep = True

            if sep is not None:
                sniff_sep = False
                dia.delimiter = sep
            # attempt to sniff the delimiter
            if sniff_sep:
                line = f.readline()
                while self.pos in self.skiprows:
                    self.pos += 1
                    line = f.readline()

                line = self._check_comments([line])[0]

                self.pos += 1
                sniffed = csv.Sniffer().sniff(line)
                dia.delimiter = sniffed.delimiter
                if self.encoding is not None:
                    self.buf.extend(list(
                        com.UnicodeReader(StringIO(line),
                                          dialect=dia,
                                          encoding=self.encoding)))
                else:
                    self.buf.extend(list(csv.reader(StringIO(line),
                                                    dialect=dia)))

            if self.encoding is not None:
                reader = com.UnicodeReader(f, dialect=dia,
                                           encoding=self.encoding,
                                           strict=True)
            else:
                reader = csv.reader(f, dialect=dia,
                                    strict=True)

        else:
            def _read():
                line = next(f)
                pat = re.compile(sep)
                if (py3compat.PY3 and isinstance(line, bytes)):
                    yield pat.split(line.decode('utf-8').strip())
                    for line in f:
                        yield pat.split(line.decode('utf-8').strip())
                else:
                    yield pat.split(line.strip())
                    for line in f:
                        yield pat.split(line.strip())
            reader = _read()

        self.data = reader

    def read(self, rows=None):
        try:
            content = self._get_lines(rows)
        except StopIteration:
            if self._first_chunk:
                content = []
            else:
                raise

        # done with first read, next time raise StopIteration
        self._first_chunk = False

        columns = list(self.orig_names)
        if len(content) == 0:  # pragma: no cover
            # DataFrame with the right metadata, even though it's length 0
            return _get_empty_meta(self.orig_names,
                                   self.index_col,
                                   self.index_names)

        # handle new style for names in index
        count_empty_content_vals = count_empty_vals(content[0])
        indexnamerow = None
        if self.has_index_names and count_empty_content_vals == len(columns):
            indexnamerow = content[0]
            content = content[1:]

        alldata = self._rows_to_cols(content)
        data = self._exclude_implicit_index(alldata)

        columns, data = self._do_date_conversions(self.columns, data)

        data = self._convert_data(data)
        index = self._make_index(data, alldata, columns)
        if indexnamerow:
            coffset = len(indexnamerow) - len(columns)
            index.names = indexnamerow[:coffset]

        return index, columns, data

    # legacy
    def get_chunk(self, size=None):
        if size is None:
            size = self.chunksize
        return self.read(nrows=size)

    def _convert_data(self, data):
        # apply converters
        clean_conv = {}

        for col, f in self.converters.iteritems():
            if isinstance(col, int) and col not in self.orig_names:
                col = self.orig_names[col]
            clean_conv[col] = f

        return self._convert_to_ndarrays(data, self.na_values, self.verbose,
                                         clean_conv)

    def _infer_columns(self):
        names = self.names

        if self.header is not None:
            if len(self.buf) > 0:
                line = self.buf[0]
            else:
                line = self._next_line()

            while self.pos <= self.header:
                line = self._next_line()

            columns = []
            for i, c in enumerate(line):
                if c == '':
                    columns.append('Unnamed: %d' % i)
                else:
                    columns.append(c)

            counts = {}
            for i, col in enumerate(columns):
                cur_count = counts.get(col, 0)
                if cur_count > 0:
                    columns[i] = '%s.%d' % (col, cur_count)
                counts[col] = cur_count + 1

            self._clear_buffer()

            if names is not None:
                if len(names) != len(columns):
                    raise Exception('Number of passed names did not match '
                                    'number of header fields in the file')
                columns = names
        else:
            if len(self.buf) > 0:
                line = self.buf[0]
            else:
                line = self._next_line()

            ncols = len(line)
            if not names:
                if self.prefix:
                    columns = ['X%d' % i for i in range(ncols)]
                else:
                    columns = range(ncols)
            else:
                columns = names

        return columns

    def _next_line(self):
        if isinstance(self.data, list):
            while self.pos in self.skiprows:
                self.pos += 1

            try:
                line = self.data[self.pos]
            except IndexError:
                raise StopIteration
        else:
            while self.pos in self.skiprows:
                next(self.data)
                self.pos += 1

            line = next(self.data)

        line = self._check_comments([line])[0]
        line = self._check_thousands([line])[0]

        self.pos += 1
        self.buf.append(line)

        return line

    def _check_comments(self, lines):
        if self.comment is None:
            return lines
        ret = []
        for l in lines:
            rl = []
            for x in l:
                if (not isinstance(x, basestring) or
                        self.comment not in x):
                    rl.append(x)
                else:
                    x = x[:x.find(self.comment)]
                    if len(x) > 0:
                        rl.append(x)
                    break
            ret.append(rl)
        return ret

    def _check_thousands(self, lines):
        if self.thousands is None:
            return lines
        nonnum = re.compile('[^-^0-9^%s^.]+' % self.thousands)
        ret = []
        for l in lines:
            rl = []
            for x in l:
                if (not isinstance(x, basestring) or
                    self.thousands not in x or
                        nonnum.search(x.strip())):
                    rl.append(x)
                else:
                    rl.append(x.replace(',', ''))
            ret.append(rl)
        return ret

    def _clear_buffer(self):
        self.buf = []

    _implicit_index = False

    def _get_index_name(self, columns):
        orig_names = list(columns)
        columns = list(columns)

        try:
            line = self._next_line()
        except StopIteration:
            line = None

        try:
            next_line = self._next_line()
        except StopIteration:
            next_line = None

        index_name = None

        # implicitly index_col=0 b/c 1 fewer column names
        implicit_first_cols = 0
        if line is not None:
            # leave it 0, #2442
            if self.index_col is not False:
                implicit_first_cols = len(line) - len(columns)

            if next_line is not None:
                if len(next_line) == len(line) + len(columns):
                    # column and index names on diff rows
                    implicit_first_cols = 0

                    self.index_col = range(len(line))
                    self.buf = self.buf[1:]

                    for c in reversed(line):
                        columns.insert(0, c)

                    return line, columns, orig_names

        if implicit_first_cols > 0:
            self._implicit_index = True
            if self.index_col is None:
                self.index_col = range(implicit_first_cols)
            index_name = None

        else:
            (index_name, columns,
             self.index_col) = _clean_index_names(columns, self.index_col)

        return index_name, orig_names, columns

    def _rows_to_cols(self, content):
        zipped_content = list(lib.to_object_array(content).T)

        col_len = len(self.orig_names)
        zip_len = len(zipped_content)

        if self._implicit_index:
            col_len += len(self.index_col)

        assert(self.skip_footer >= 0)

        if col_len != zip_len and self.index_col is not False:
            row_num = -1
            i = 0
            for (i, l) in enumerate(content):
                if len(l) != col_len:
                    break

            footers = 0
            if self.skip_footer:
                footers = self.skip_footer

            row_num = self.pos - (len(content) - i + footers)

            msg = ('Expected %d fields in line %d, saw %d' %
                   (col_len, row_num + 1, zip_len))
            raise ValueError(msg)

        return zipped_content

    def _get_lines(self, rows=None):
        source = self.data
        lines = self.buf

        # already fetched some number
        if rows is not None:
            rows -= len(self.buf)

        if isinstance(source, list):
            if self.pos > len(source):
                raise StopIteration
            if rows is None:
                lines.extend(source[self.pos:])
                self.pos = len(source)
            else:
                lines.extend(source[self.pos:self.pos + rows])
                self.pos += rows
        else:
            new_rows = []
            try:
                if rows is not None:
                    for _ in xrange(rows):
                        new_rows.append(next(source))
                    lines.extend(new_rows)
                else:
                    rows = 0
                    while True:
                        try:
                            new_rows.append(next(source))
                            rows += 1
                        except csv.Error, inst:
                            if 'newline inside string' in str(inst):
                                row_num = str(self.pos + rows)
                                msg = ('EOF inside string starting with line '
                                       + row_num)
                                raise Exception(msg)
                            raise
            except StopIteration:
                lines.extend(new_rows)
                if len(lines) == 0:
                    raise
            self.pos += len(new_rows)

        self.buf = []

        if self.skip_footer:
            lines = lines[:-self.skip_footer]

        lines = self._check_comments(lines)
        return self._check_thousands(lines)


def _make_date_converter(date_parser=None, dayfirst=False):
    def converter(*date_cols):
        if date_parser is None:
            strs = _concat_date_cols(date_cols)
            try:
                return tslib.array_to_datetime(com._ensure_object(strs),
                                               utc=None, dayfirst=dayfirst)
            except:
                return lib.try_parse_dates(strs, dayfirst=dayfirst)
        else:
            try:
                result = date_parser(*date_cols)
                if isinstance(result, datetime.datetime):
                    raise Exception('scalar parser')
                return result
            except Exception:
                try:
                    return lib.try_parse_dates(_concat_date_cols(date_cols),
                                               parser=date_parser,
                                               dayfirst=dayfirst)
                except Exception:
                    return generic_parser(date_parser, *date_cols)

    return converter


def _process_date_conversion(data_dict, converter, parse_spec,
                             index_col, index_names, columns,
                             keep_date_col=False):
    def _isindex(colspec):
        return ((isinstance(index_col, list) and
                 colspec in index_col)
                or (isinstance(index_names, list) and
                    colspec in index_names))

    new_cols = []
    new_data = {}

    orig_names = columns
    columns = list(columns)

    date_cols = set()

    if parse_spec is None or isinstance(parse_spec, bool):
        return data_dict, columns

    if isinstance(parse_spec, list):
        # list of column lists
        for colspec in parse_spec:
            if np.isscalar(colspec):
                if isinstance(colspec, int) and colspec not in data_dict:
                    colspec = orig_names[colspec]
                if _isindex(colspec):
                    continue
                data_dict[colspec] = converter(data_dict[colspec])
            else:
                new_name, col, old_names = _try_convert_dates(
                    converter, colspec, data_dict, orig_names)
                if new_name in data_dict:
                    raise ValueError('New date column already in dict %s' %
                                     new_name)
                new_data[new_name] = col
                new_cols.append(new_name)
                date_cols.update(old_names)

    elif isinstance(parse_spec, dict):
        # dict of new name to column list
        for new_name, colspec in parse_spec.iteritems():
            if new_name in data_dict:
                raise ValueError('Date column %s already in dict' %
                                 new_name)

            _, col, old_names = _try_convert_dates(converter, colspec,
                                                   data_dict, orig_names)

            new_data[new_name] = col
            new_cols.append(new_name)
            date_cols.update(old_names)

    data_dict.update(new_data)
    new_cols.extend(columns)

    if not keep_date_col:
        for c in list(date_cols):
            data_dict.pop(c)
            new_cols.remove(c)

    return data_dict, new_cols


def _try_convert_dates(parser, colspec, data_dict, columns):
    colset = set(columns)
    colnames = []

    for c in colspec:
        if c in colset:
            colnames.append(c)
        elif isinstance(c, int) and c not in columns:
            colnames.append(str(columns[c]))
        else:
            colnames.append(c)

    new_name = '_'.join([str(x) for x in colnames])
    to_parse = [data_dict[c] for c in colnames if c in data_dict]

    try:
        new_col = parser(*to_parse)
    except DateConversionError:
        new_col = parser(_concat_date_cols(to_parse))
    return new_name, new_col, colnames


def _clean_na_values(na_values, keep_default_na=True):
    if na_values is None and keep_default_na:
        na_values = _NA_VALUES
    elif isinstance(na_values, dict):
        if keep_default_na:
            for k, v in na_values.iteritems():
                v = set(list(v)) | _NA_VALUES
                na_values[k] = v
    else:
        if not com.is_list_like(na_values):
            na_values = [na_values]
        na_values = set(list(na_values))
        if keep_default_na:
            na_values = na_values | _NA_VALUES

    return na_values


def _clean_index_names(columns, index_col):
    if not _is_index_col(index_col):
        return None, columns, index_col

    columns = list(columns)

    cp_cols = list(columns)
    index_names = []

    # don't mutate
    index_col = list(index_col)

    for i, c in enumerate(index_col):
        if isinstance(c, basestring):
            index_names.append(c)
            for j, name in enumerate(cp_cols):
                if name == c:
                    index_col[i] = j
                    columns.remove(name)
                    break
        else:
            name = cp_cols[c]
            columns.remove(name)
            index_names.append(name)

    # hack
    if isinstance(index_names[0], basestring) and 'Unnamed' in index_names[0]:
        index_names[0] = None

    return index_names, columns, index_col


def _get_empty_meta(columns, index_col, index_names):
    columns = list(columns)

    if index_col is not None:
        index = MultiIndex.from_arrays([[]] * len(index_col),
                                       names=index_names)
        for n in index_col:
            columns.pop(n)
    else:
        index = Index([])

    return index, columns, {}


def _get_na_values(col, na_values):
    if isinstance(na_values, dict):
        if col in na_values:
            return set(list(na_values[col]))
        else:
            return _NA_VALUES
    else:
        return na_values


def _get_col_names(colspec, columns):
    colset = set(columns)
    colnames = []
    for c in colspec:
        if c in colset:
            colnames.append(c)
        elif isinstance(c, int):
            colnames.append(columns[c])
    return colnames


def _concat_date_cols(date_cols):
    if len(date_cols) == 1:
        if py3compat.PY3:
            return np.array([unicode(x) for x in date_cols[0]], dtype=object)
        else:
            return np.array([str(x) if not isinstance(x, basestring) else x
                             for x in date_cols[0]], dtype=object)

    # stripped = [map(str.strip, x) for x in date_cols]
    rs = np.array([' '.join([unicode(y) for y in x])
                   for x in zip(*date_cols)], dtype=object)
    return rs


class FixedWidthReader(object):
    """
    A reader of fixed-width lines.
    """
    def __init__(self, f, colspecs, filler, thousands=None):
        self.f = f
        self.colspecs = colspecs
        self.filler = filler  # Empty characters between fields.
        self.thousands = thousands

        assert isinstance(colspecs, (tuple, list))
        for colspec in colspecs:
            assert isinstance(colspec, (tuple, list))
            assert len(colspec) == 2
            assert isinstance(colspec[0], int)
            assert isinstance(colspec[1], int)

    def next(self):
        line = next(self.f)
        # Note: 'colspecs' is a sequence of half-open intervals.
        return [line[fromm:to].strip(self.filler or ' ')
                for (fromm, to) in self.colspecs]

    # Iterator protocol in Python 3 uses __next__()
    __next__ = next


class FixedWidthFieldParser(PythonParser):
    """
    Specialization that Converts fixed-width fields into DataFrames.
    See PythonParser for details.
    """
    def __init__(self, f, **kwds):
        # Support iterators, convert to a list.
        self.colspecs = list(kwds.pop('colspecs'))

        PythonParser.__init__(self, f, **kwds)

    def _make_reader(self, f):
        self.data = FixedWidthReader(f, self.colspecs, self.delimiter)


#----------------------------------------------------------------------
# ExcelFile class

_openpyxl_msg = ("\nFor parsing .xlsx files 'openpyxl' is required.\n"
                 "You can install it via 'easy_install openpyxl' or "
                 "'pip install openpyxl'.\nAlternatively, you could save"
                 " the .xlsx file as a .xls file.\n")


class ExcelFile(object):
    """
    Class for parsing tabular excel sheets into DataFrame objects.
    Uses xlrd for parsing .xls files or openpyxl for .xlsx files.
    See ExcelFile.parse for more documentation

    Parameters
    ----------
    path : string or file-like object
        Path to xls or xlsx file
    kind : {'xls', 'xlsx', None}, default None
    """
    def __init__(self, path_or_buf, kind=None):
        self.kind = kind
        self.use_xlsx = kind == 'xls'

        self.path_or_buf = path_or_buf
        self.tmpfile = None

        if isinstance(path_or_buf, basestring):
            if kind == 'xls' or (kind is None and
                                 path_or_buf.endswith('.xls')):
                self.use_xlsx = False
                import xlrd
                self.book = xlrd.open_workbook(path_or_buf)
            else:
                self.use_xlsx = True
                try:
                    from openpyxl.reader.excel import load_workbook
                    self.book = load_workbook(path_or_buf, use_iterators=True)
                except ImportError:  # pragma: no cover
                    raise ImportError(_openpyxl_msg)
        else:
            data = path_or_buf.read()

            if self.kind == 'xls':
                import xlrd
                self.book = xlrd.open_workbook(file_contents=data)
            elif self.kind == 'xlsx':
                from openpyxl.reader.excel import load_workbook
                buf = py3compat.BytesIO(data)
                self.book = load_workbook(buf, use_iterators=True)
            else:
                try:
                    import xlrd
                    self.book = xlrd.open_workbook(file_contents=data)
                    self.use_xlsx = False
                except Exception:
                    self.use_xlsx = True
                    from openpyxl.reader.excel import load_workbook
                    buf = py3compat.BytesIO(data)
                    self.book = load_workbook(buf, use_iterators=True)

    def __repr__(self):
        return object.__repr__(self)

    def parse(self, sheetname, header=0, skiprows=None, skip_footer=0,
              index_col=None, parse_cols=None, parse_dates=False,
              date_parser=None, na_values=None, thousands=None, chunksize=None,
              **kwds):
        """
        Read Excel table into DataFrame

        Parameters
        ----------
        sheetname : string
            Name of Excel sheet
        header : int, default 0
            Row to use for the column labels of the parsed DataFrame
        skiprows : list-like
            Rows to skip at the beginning (0-indexed)
        skip_footer : int, default 0
            Rows at the end to skip (0-indexed)
        index_col : int, default None
            Column to use as the row labels of the DataFrame. Pass None if
            there is no such column
        parse_cols : int or list, default None
            If None then parse all columns,
            If int then indicates last column to be parsed
            If list of ints then indicates list of column numbers to be parsed
            If string then indicates comma separated list of column names and
                column ranges (e.g. "A:E" or "A,C,E:F")
        na_values : list-like, default None
            List of additional strings to recognize as NA/NaN

        Returns
        -------
        parsed : DataFrame
        """

        # has_index_names: boolean, default False
        #     True if the cols defined in index_col have an index name and are
        #     not in the header
        has_index_names = False  # removed as new argument of API function

        skipfooter = kwds.pop('skipfooter', None)
        if skipfooter is not None:
            skip_footer = skipfooter

        choose = {True: self._parse_xlsx,
                  False: self._parse_xls}
        return choose[self.use_xlsx](sheetname, header=header,
                                     skiprows=skiprows, index_col=index_col,
                                     has_index_names=has_index_names,
                                     parse_cols=parse_cols,
                                     parse_dates=parse_dates,
                                     date_parser=date_parser,
                                     na_values=na_values,
                                     thousands=thousands,
                                     chunksize=chunksize,
                                     skip_footer=skip_footer)

    def _should_parse(self, i, parse_cols):

        def _range2cols(areas):
            """
            Convert comma separated list of column names and column ranges to a
            list of 0-based column indexes.

            >>> _range2cols('A:E')
            [0, 1, 2, 3, 4]
            >>> _range2cols('A,C,Z:AB')
            [0, 2, 25, 26, 27]
            """
            def _excel2num(x):
                "Convert Excel column name like 'AB' to 0-based column index"
                return reduce(lambda s, a: s * 26 + ord(a) - ord('A') + 1, x.upper().strip(), 0) - 1

            cols = []
            for rng in areas.split(','):
                if ':' in rng:
                    rng = rng.split(':')
                    cols += range(_excel2num(rng[0]), _excel2num(rng[1]) + 1)
                else:
                    cols.append(_excel2num(rng))
            return cols

        if isinstance(parse_cols, int):
            return i <= parse_cols
        elif isinstance(parse_cols, basestring):
            return i in _range2cols(parse_cols)
        else:
            return i in parse_cols

    def _parse_xlsx(self, sheetname, header=0, skiprows=None,
                    skip_footer=0, index_col=None, has_index_names=False,
                    parse_cols=None, parse_dates=False, date_parser=None,
                    na_values=None, thousands=None, chunksize=None):
        sheet = self.book.get_sheet_by_name(name=sheetname)
        data = []

        # it brings a new method: iter_rows()
        should_parse = {}

        for row in sheet.iter_rows():
            row_data = []
            for j, cell in enumerate(row):

                if parse_cols is not None and j not in should_parse:
                    should_parse[j] = self._should_parse(j, parse_cols)

                if parse_cols is None or should_parse[j]:
                    row_data.append(cell.internal_value)
            data.append(row_data)

        if header is not None:
            data[header] = _trim_excel_header(data[header])

        parser = TextParser(data, header=header, index_col=index_col,
                            has_index_names=has_index_names,
                            na_values=na_values,
                            thousands=thousands,
                            parse_dates=parse_dates,
                            date_parser=date_parser,
                            skiprows=skiprows,
                            skip_footer=skip_footer,
                            chunksize=chunksize)

        return parser.read()

    def _parse_xls(self, sheetname, header=0, skiprows=None,
                   skip_footer=0, index_col=None, has_index_names=None,
                   parse_cols=None, parse_dates=False, date_parser=None,
                   na_values=None, thousands=None, chunksize=None):
        from xlrd import xldate_as_tuple, XL_CELL_DATE, XL_CELL_ERROR

        datemode = self.book.datemode
        sheet = self.book.sheet_by_name(sheetname)

        data = []
        should_parse = {}
        for i in range(sheet.nrows):
            row = []
            for j, (value, typ) in enumerate(izip(sheet.row_values(i),
                                                  sheet.row_types(i))):
                if parse_cols is not None and j not in should_parse:
                    should_parse[j] = self._should_parse(j, parse_cols)

                if parse_cols is None or should_parse[j]:
                    if typ == XL_CELL_DATE:
                        dt = xldate_as_tuple(value, datemode)
                        # how to produce this first case?
                        if dt[0] < datetime.MINYEAR:  # pragma: no cover
                            value = datetime.time(*dt[3:])
                        else:
                            value = datetime.datetime(*dt)
                    if typ == XL_CELL_ERROR:
                        value = np.nan
                    row.append(value)
            data.append(row)

        if header is not None:
            data[header] = _trim_excel_header(data[header])

        parser = TextParser(data, header=header, index_col=index_col,
                            has_index_names=has_index_names,
                            na_values=na_values,
                            thousands=thousands,
                            parse_dates=parse_dates,
                            date_parser=date_parser,
                            skiprows=skiprows,
                            skip_footer=skip_footer,
                            chunksize=chunksize)

        return parser.read()

    @property
    def sheet_names(self):
        if self.use_xlsx:
            return self.book.get_sheet_names()
        else:
            return self.book.sheet_names()


def _trim_excel_header(row):
    # trim header row so auto-index inference works
    # xlrd uses '' , openpyxl None
    while len(row) > 0 and (row[0] == '' or row[0] is None):
        row = row[1:]
    return row


class CellStyleConverter(object):
    """
    Utility Class which converts a style dict to xlrd or openpyxl style
    """

    @staticmethod
    def to_xls(style_dict, num_format_str=None):
        """
        converts a style_dict to an xlwt style object
        Parameters
        ----------
        style_dict: style dictionary to convert
        """
        import xlwt

        def style_to_xlwt(item, firstlevel=True, field_sep=',', line_sep=';'):
            """helper wich recursively generate an xlwt easy style string
            for example:

              hstyle = {"font": {"bold": True},
              "border": {"top": "thin",
                        "right": "thin",
                        "bottom": "thin",
                        "left": "thin"},
              "align": {"horiz": "center"}}
              will be converted to
              font: bold on; \
                      border: top thin, right thin, bottom thin, left thin; \
                      align: horiz center;
            """
            if hasattr(item, 'items'):
                if firstlevel:
                    it = ["%s: %s" % (key, style_to_xlwt(value, False))
                          for key, value in item.items()]
                    out = "%s " % (line_sep).join(it)
                    return out
                else:
                    it = ["%s %s" % (key, style_to_xlwt(value, False))
                          for key, value in item.items()]
                    out = "%s " % (field_sep).join(it)
                    return out
            else:
                item = "%s" % item
                item = item.replace("True", "on")
                item = item.replace("False", "off")
                return item

        if style_dict:
            xlwt_stylestr = style_to_xlwt(style_dict)
            style = xlwt.easyxf(xlwt_stylestr, field_sep=',', line_sep=';')
        else:
            style = xlwt.XFStyle()
        if num_format_str is not None:
            style.num_format_str = num_format_str

        return style

    @staticmethod
    def to_xlsx(style_dict):
        """
        converts a style_dict to an openpyxl style object
        Parameters
        ----------
        style_dict: style dictionary to convert
        """

        from openpyxl.style import Style
        xls_style = Style()
        for key, value in style_dict.items():
            for nk, nv in value.items():
                if key == "borders":
                    (xls_style.borders.__getattribute__(nk)
                     .__setattr__('border_style', nv))
                else:
                    xls_style.__getattribute__(key).__setattr__(nk, nv)

        return xls_style


def _conv_value(val):
    # convert value for excel dump
    if isinstance(val, np.int64):
        val = int(val)
    elif isinstance(val, np.bool8):
        val = bool(val)
    elif isinstance(val, Period):
        val = "%s" % val

    return val


class ExcelWriter(object):
    """
    Class for writing DataFrame objects into excel sheets, uses xlwt for xls,
    openpyxl for xlsx.  See DataFrame.to_excel for typical usage.

    Parameters
    ----------
    path : string
        Path to xls file
    """
    def __init__(self, path):
        self.use_xlsx = True
        if path.endswith('.xls'):
            self.use_xlsx = False
            import xlwt
            self.book = xlwt.Workbook()
            self.fm_datetime = xlwt.easyxf(
                num_format_str='YYYY-MM-DD HH:MM:SS')
            self.fm_date = xlwt.easyxf(num_format_str='YYYY-MM-DD')
        else:
            from openpyxl.workbook import Workbook
            self.book = Workbook()  # optimized_write=True)
            # open pyxl 1.6.1 adds a dummy sheet remove it
            if self.book.worksheets:
                self.book.remove_sheet(self.book.worksheets[0])
        self.path = path
        self.sheets = {}
        self.cur_sheet = None

    def save(self):
        """
        Save workbook to disk
        """
        self.book.save(self.path)

    def write_cells(self, cells, sheet_name=None, startrow=0, startcol=0):
        """
        Write given formated cells into Excel an excel sheet

        Parameters
        ----------
        cells : generator
            cell of formated data to save to Excel sheet
        sheet_name : string, default None
            Name of Excel sheet, if None, then use self.cur_sheet
        startrow: upper left cell row to dump data frame
        startcol: upper left cell column to dump data frame
        """
        if sheet_name is None:
            sheet_name = self.cur_sheet
        if sheet_name is None:  # pragma: no cover
            raise Exception('Must pass explicit sheet_name or set '
                            'cur_sheet property')
        if self.use_xlsx:
            self._writecells_xlsx(cells, sheet_name, startrow, startcol)
        else:
            self._writecells_xls(cells, sheet_name, startrow, startcol)

    def _writecells_xlsx(self, cells, sheet_name, startrow, startcol):

        from openpyxl.cell import get_column_letter

        if sheet_name in self.sheets:
            wks = self.sheets[sheet_name]
        else:
            wks = self.book.create_sheet()
            wks.title = sheet_name
            self.sheets[sheet_name] = wks

        for cell in cells:
            colletter = get_column_letter(startcol + cell.col + 1)
            xcell = wks.cell("%s%s" % (colletter, startrow + cell.row + 1))
            xcell.value = _conv_value(cell.val)
            if cell.style:
                style = CellStyleConverter.to_xlsx(cell.style)
                for field in style.__fields__:
                    xcell.style.__setattr__(field,
                                            style.__getattribute__(field))

            if isinstance(cell.val, datetime.datetime):
                xcell.style.number_format.format_code = "YYYY-MM-DD HH:MM:SS"
            elif isinstance(cell.val, datetime.date):
                xcell.style.number_format.format_code = "YYYY-MM-DD"

            # merging requires openpyxl latest (works on 1.6.1)
            # todo add version check
            if cell.mergestart is not None and cell.mergeend is not None:
                cletterstart = get_column_letter(startcol + cell.col + 1)
                cletterend = get_column_letter(startcol + cell.mergeend + 1)

                wks.merge_cells('%s%s:%s%s' % (cletterstart,
                                               startrow + cell.row + 1,
                                               cletterend,
                                               startrow + cell.mergestart + 1))

    def _writecells_xls(self, cells, sheet_name, startrow, startcol):
        if sheet_name in self.sheets:
            wks = self.sheets[sheet_name]
        else:
            wks = self.book.add_sheet(sheet_name)
            self.sheets[sheet_name] = wks

        style_dict = {}

        for cell in cells:
            val = _conv_value(cell.val)

            num_format_str = None
            if isinstance(cell.val, datetime.datetime):
                num_format_str = "YYYY-MM-DD HH:MM:SS"
            if isinstance(cell.val, datetime.date):
                num_format_str = "YYYY-MM-DD"

            stylekey = json.dumps(cell.style)
            if num_format_str:
                stylekey += num_format_str

            if stylekey in style_dict:
                style = style_dict[stylekey]
            else:
                style = CellStyleConverter.to_xls(cell.style, num_format_str)
                style_dict[stylekey] = style

            if cell.mergestart is not None and cell.mergeend is not None:
                wks.write_merge(startrow + cell.row,
                                startrow + cell.mergestart,
                                startcol + cell.col,
                                startcol + cell.mergeend,
                                val, style)
            else:
                wks.write(startrow + cell.row,
                          startcol + cell.col,
                          val, style)
