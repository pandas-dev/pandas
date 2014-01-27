# pylint: disable=E1101

from pandas.compat import u, range, map
from datetime import datetime, date
import os

import nose

from numpy import nan
import numpy as np

from pandas import DataFrame, Index, MultiIndex
from pandas.io.parsers import read_csv
from pandas.io.excel import (
    ExcelFile, ExcelWriter, read_excel, _XlwtWriter, _OpenpyxlWriter,
    register_writer, _XlsxWriter
)
from pandas.util.testing import ensure_clean
from pandas.core.config import set_option, get_option
import pandas.util.testing as tm
import pandas as pd


def _skip_if_no_xlrd():
    try:
        import xlrd
        ver = tuple(map(int, xlrd.__VERSION__.split(".")[:2]))
        if ver < (0, 9):
            raise nose.SkipTest('xlrd < 0.9, skipping')
    except ImportError:
        raise nose.SkipTest('xlrd not installed, skipping')


def _skip_if_no_xlwt():
    try:
        import xlwt  # NOQA
    except ImportError:
        raise nose.SkipTest('xlwt not installed, skipping')


def _skip_if_no_openpyxl():
    try:
        import openpyxl  # NOQA
    except ImportError:
        raise nose.SkipTest('openpyxl not installed, skipping')


def _skip_if_no_xlsxwriter():
    try:
        import xlsxwriter  # NOQA
    except ImportError:
        raise nose.SkipTest('xlsxwriter not installed, skipping')


def _skip_if_no_excelsuite():
    _skip_if_no_xlrd()
    _skip_if_no_xlwt()
    _skip_if_no_openpyxl()


_seriesd = tm.getSeriesData()
_tsd = tm.getTimeSeriesData()
_frame = DataFrame(_seriesd)[:10]
_frame2 = DataFrame(_seriesd, columns=['D', 'C', 'B', 'A'])[:10]
_tsframe = tm.makeTimeDataFrame()[:5]
_mixed_frame = _frame.copy()
_mixed_frame['foo'] = 'bar'


class SharedItems(object):
    def setUp(self):
        self.dirpath = tm.get_data_path()
        self.csv1 = os.path.join(self.dirpath, 'test1.csv')
        self.csv2 = os.path.join(self.dirpath, 'test2.csv')
        self.xls1 = os.path.join(self.dirpath, 'test.xls')
        self.xlsx1 = os.path.join(self.dirpath, 'test.xlsx')
        self.frame = _frame.copy()
        self.frame2 = _frame2.copy()
        self.tsframe = _tsframe.copy()
        self.mixed_frame = _mixed_frame.copy()

    def read_csv(self, *args, **kwds):
        kwds = kwds.copy()
        kwds['engine'] = 'python'
        return read_csv(*args, **kwds)


class ExcelReaderTests(SharedItems, tm.TestCase):
    def test_parse_cols_int(self):
        _skip_if_no_openpyxl()
        _skip_if_no_xlrd()

        suffix = ['xls', 'xlsx', 'xlsm']

        for s in suffix:
            pth = os.path.join(self.dirpath, 'test.%s' % s)
            xls = ExcelFile(pth)
            df = xls.parse('Sheet1', index_col=0, parse_dates=True,
                           parse_cols=3)
            df2 = self.read_csv(self.csv1, index_col=0, parse_dates=True)
            df2 = df2.reindex(columns=['A', 'B', 'C'])
            df3 = xls.parse('Sheet2', skiprows=[1], index_col=0,
                            parse_dates=True, parse_cols=3)
            # TODO add index to xls file)
            tm.assert_frame_equal(df, df2, check_names=False)
            tm.assert_frame_equal(df3, df2, check_names=False)

    def test_parse_cols_list(self):
        _skip_if_no_openpyxl()
        _skip_if_no_xlrd()

        suffix = ['xls', 'xlsx', 'xlsm']

        for s in suffix:
            pth = os.path.join(self.dirpath, 'test.%s' % s)
            xls = ExcelFile(pth)
            df = xls.parse('Sheet1', index_col=0, parse_dates=True,
                           parse_cols=[0, 2, 3])
            df2 = self.read_csv(self.csv1, index_col=0, parse_dates=True)
            df2 = df2.reindex(columns=['B', 'C'])
            df3 = xls.parse('Sheet2', skiprows=[1], index_col=0,
                            parse_dates=True,
                            parse_cols=[0, 2, 3])
            # TODO add index to xls file)
            tm.assert_frame_equal(df, df2, check_names=False)
            tm.assert_frame_equal(df3, df2, check_names=False)

    def test_parse_cols_str(self):
        _skip_if_no_openpyxl()
        _skip_if_no_xlrd()

        suffix = ['xls', 'xlsx', 'xlsm']

        for s in suffix:

            pth = os.path.join(self.dirpath, 'test.%s' % s)
            xls = ExcelFile(pth)

            df = xls.parse('Sheet1', index_col=0, parse_dates=True,
                           parse_cols='A:D')
            df2 = read_csv(self.csv1, index_col=0, parse_dates=True)
            df2 = df2.reindex(columns=['A', 'B', 'C'])
            df3 = xls.parse('Sheet2', skiprows=[1], index_col=0,
                            parse_dates=True, parse_cols='A:D')
            # TODO add index to xls, read xls ignores index name ?
            tm.assert_frame_equal(df, df2, check_names=False)
            tm.assert_frame_equal(df3, df2, check_names=False)
            del df, df2, df3

            df = xls.parse('Sheet1', index_col=0, parse_dates=True,
                           parse_cols='A,C,D')
            df2 = read_csv(self.csv1, index_col=0, parse_dates=True)
            df2 = df2.reindex(columns=['B', 'C'])
            df3 = xls.parse('Sheet2', skiprows=[1], index_col=0,
                            parse_dates=True,
                            parse_cols='A,C,D')
            # TODO add index to xls file
            tm.assert_frame_equal(df, df2, check_names=False)
            tm.assert_frame_equal(df3, df2, check_names=False)
            del df, df2, df3

            df = xls.parse('Sheet1', index_col=0, parse_dates=True,
                           parse_cols='A,C:D')
            df2 = read_csv(self.csv1, index_col=0, parse_dates=True)
            df2 = df2.reindex(columns=['B', 'C'])
            df3 = xls.parse('Sheet2', skiprows=[1], index_col=0,
                            parse_dates=True,
                            parse_cols='A,C:D')
            tm.assert_frame_equal(df, df2, check_names=False)
            tm.assert_frame_equal(df3, df2, check_names=False)

    def test_excel_stop_iterator(self):
        _skip_if_no_xlrd()

        excel_data = ExcelFile(os.path.join(self.dirpath, 'test2.xls'))
        parsed = excel_data.parse('Sheet1')
        expected = DataFrame([['aaaa', 'bbbbb']], columns=['Test', 'Test1'])
        tm.assert_frame_equal(parsed, expected)

    def test_excel_cell_error_na(self):
        _skip_if_no_xlrd()

        excel_data = ExcelFile(os.path.join(self.dirpath, 'test3.xls'))
        parsed = excel_data.parse('Sheet1')
        expected = DataFrame([[np.nan]], columns=['Test'])
        tm.assert_frame_equal(parsed, expected)

    def test_excel_passes_na(self):
        _skip_if_no_xlrd()

        excel_data = ExcelFile(os.path.join(self.dirpath, 'test2.xlsx'))
        parsed = excel_data.parse('Sheet1', keep_default_na=False,
                                  na_values=['apple'])
        expected = DataFrame([['NA'], [1], ['NA'], [np.nan], ['rabbit']],
                             columns=['Test'])
        tm.assert_frame_equal(parsed, expected)

        parsed = excel_data.parse('Sheet1', keep_default_na=True,
                                  na_values=['apple'])
        expected = DataFrame([[np.nan], [1], [np.nan], [np.nan], ['rabbit']],
                             columns=['Test'])
        tm.assert_frame_equal(parsed, expected)

    def check_excel_table_sheet_by_index(self, filename, csvfile):
        import xlrd

        pth = os.path.join(self.dirpath, filename)
        xls = ExcelFile(pth)
        df = xls.parse(0, index_col=0, parse_dates=True)
        df2 = self.read_csv(csvfile, index_col=0, parse_dates=True)
        df3 = xls.parse(1, skiprows=[1], index_col=0, parse_dates=True)
        tm.assert_frame_equal(df, df2, check_names=False)
        tm.assert_frame_equal(df3, df2, check_names=False)

        df4 = xls.parse(0, index_col=0, parse_dates=True, skipfooter=1)
        df5 = xls.parse(0, index_col=0, parse_dates=True, skip_footer=1)
        tm.assert_frame_equal(df4, df.ix[:-1])
        tm.assert_frame_equal(df4, df5)

        self.assertRaises(xlrd.XLRDError, xls.parse, 'asdf')

    def test_excel_table_sheet_by_index(self):
        _skip_if_no_xlrd()
        for filename, csvfile in [(self.xls1, self.csv1),
                                  (self.xlsx1, self.csv1)]:
            self.check_excel_table_sheet_by_index(filename, csvfile)

    def test_excel_table(self):
        _skip_if_no_xlrd()

        pth = os.path.join(self.dirpath, 'test.xls')
        xls = ExcelFile(pth)
        df = xls.parse('Sheet1', index_col=0, parse_dates=True)
        df2 = self.read_csv(self.csv1, index_col=0, parse_dates=True)
        df3 = xls.parse('Sheet2', skiprows=[1], index_col=0, parse_dates=True)
        tm.assert_frame_equal(df, df2, check_names=False)
        tm.assert_frame_equal(df3, df2, check_names=False)

        df4 = xls.parse('Sheet1', index_col=0, parse_dates=True,
                        skipfooter=1)
        df5 = xls.parse('Sheet1', index_col=0, parse_dates=True,
                        skip_footer=1)
        tm.assert_frame_equal(df4, df.ix[:-1])
        tm.assert_frame_equal(df4, df5)

    def test_excel_read_buffer(self):
        _skip_if_no_xlrd()
        _skip_if_no_openpyxl()

        pth = os.path.join(self.dirpath, 'test.xls')
        f = open(pth, 'rb')
        xls = ExcelFile(f)
        # it works
        xls.parse('Sheet1', index_col=0, parse_dates=True)

        pth = os.path.join(self.dirpath, 'test.xlsx')
        f = open(pth, 'rb')
        xl = ExcelFile(f)
        xl.parse('Sheet1', index_col=0, parse_dates=True)

    def test_read_xlrd_Book(self):
        _skip_if_no_xlrd()
        _skip_if_no_xlwt()

        import xlrd

        df = self.frame

        with ensure_clean('.xls') as pth:
            df.to_excel(pth, "SheetA")
            book = xlrd.open_workbook(pth)

            with ExcelFile(book, engine="xlrd") as xl:
                result = xl.parse("SheetA")
                tm.assert_frame_equal(df, result)

            result = read_excel(book, sheetname="SheetA", engine="xlrd")
            tm.assert_frame_equal(df, result)

    def test_xlsx_table(self):
        _skip_if_no_xlrd()
        _skip_if_no_openpyxl()

        pth = os.path.join(self.dirpath, 'test.xlsx')
        xlsx = ExcelFile(pth)
        df = xlsx.parse('Sheet1', index_col=0, parse_dates=True)
        df2 = self.read_csv(self.csv1, index_col=0, parse_dates=True)
        df3 = xlsx.parse('Sheet2', skiprows=[1], index_col=0, parse_dates=True)

        # TODO add index to xlsx file
        tm.assert_frame_equal(df, df2, check_names=False)
        tm.assert_frame_equal(df3, df2, check_names=False)

        df4 = xlsx.parse('Sheet1', index_col=0, parse_dates=True,
                         skipfooter=1)
        df5 = xlsx.parse('Sheet1', index_col=0, parse_dates=True,
                         skip_footer=1)
        tm.assert_frame_equal(df4, df.ix[:-1])
        tm.assert_frame_equal(df4, df5)

    def test_reader_closes_file(self):
        _skip_if_no_xlrd()
        _skip_if_no_openpyxl()

        pth = os.path.join(self.dirpath, 'test.xlsx')
        f = open(pth, 'rb')
        with ExcelFile(f) as xlsx:
            # parses okay
            xlsx.parse('Sheet1', index_col=0)

        self.assertTrue(f.closed)

    def test_reader_special_dtypes(self):
        _skip_if_no_xlrd()

        expected = DataFrame.from_items([
            ("IntCol", [1, 2, -3, 4, 0]),
            ("FloatCol", [1.25, 2.25, 1.83, 1.92, 0.0000000005]),
            ("BoolCol", [True, False, True, True, False]),
            ("StrCol", [1, 2, 3, 4, 5]),
            # GH5394 - this is why convert_float isn't vectorized
            ("Str2Col", ["a", 3, "c", "d", "e"]),
            ("DateCol", [datetime(2013, 10, 30), datetime(2013, 10, 31),
                         datetime(1905, 1, 1), datetime(2013, 12, 14),
                         datetime(2015, 3, 14)])
        ])

        xlsx_path = os.path.join(self.dirpath, 'test_types.xlsx')
        xls_path = os.path.join(self.dirpath, 'test_types.xls')

        # should read in correctly and infer types
        for path in (xls_path, xlsx_path):
            actual = read_excel(path, 'Sheet1')
            tm.assert_frame_equal(actual, expected)

        # if not coercing number, then int comes in as float
        float_expected = expected.copy()
        float_expected["IntCol"] = float_expected["IntCol"].astype(float)
        float_expected.loc[1, "Str2Col"] = 3.0
        for path in (xls_path, xlsx_path):
            actual = read_excel(path, 'Sheet1', convert_float=False)
            tm.assert_frame_equal(actual, float_expected)

        # check setting Index (assuming xls and xlsx are the same here)
        for icol, name in enumerate(expected.columns):
            actual = read_excel(xlsx_path, 'Sheet1', index_col=icol)
            actual2 = read_excel(xlsx_path, 'Sheet1', index_col=name)
            exp = expected.set_index(name)
            tm.assert_frame_equal(actual, exp)
            tm.assert_frame_equal(actual2, exp)

        # convert_float and converters should be different but both accepted
        expected["StrCol"] = expected["StrCol"].apply(str)
        actual = read_excel(xlsx_path, 'Sheet1', converters={"StrCol": str})
        tm.assert_frame_equal(actual, expected)

        no_convert_float = float_expected.copy()
        no_convert_float["StrCol"] = no_convert_float["StrCol"].apply(str)
        actual = read_excel(xlsx_path, 'Sheet1', converters={"StrCol": str},
                           convert_float=False)
        tm.assert_frame_equal(actual, no_convert_float)


class ExcelWriterBase(SharedItems):
    # Base class for test cases to run with different Excel writers.
    # To add a writer test, define the following:
    # 1. A check_skip function that skips your tests if your writer isn't
    #    installed.
    # 2. Add a property ext, which is the file extension that your writer
    #    writes to. (needs to start with '.' so it's a valid path)
    # 3. Add a property engine_name, which is the name of the writer class.

    # Test with MultiIndex and Hierarchical Rows as merged cells.
    merge_cells = True

    def setUp(self):
        self.check_skip()
        super(ExcelWriterBase, self).setUp()
        self.option_name = 'io.excel.%s.writer' % self.ext.strip('.')
        self.prev_engine = get_option(self.option_name)
        set_option(self.option_name, self.engine_name)

    def tearDown(self):
        set_option(self.option_name, self.prev_engine)

    def test_excel_sheet_by_name_raise(self):
        _skip_if_no_xlrd()
        import xlrd

        with ensure_clean(self.ext) as pth:
            gt = DataFrame(np.random.randn(10, 2))
            gt.to_excel(pth)
            xl = ExcelFile(pth)
            df = xl.parse(0)
            tm.assert_frame_equal(gt, df)

            self.assertRaises(xlrd.XLRDError, xl.parse, '0')

    def test_excelwriter_contextmanager(self):
        _skip_if_no_xlrd()

        with ensure_clean(self.ext) as pth:
            with ExcelWriter(pth) as writer:
                self.frame.to_excel(writer, 'Data1')
                self.frame2.to_excel(writer, 'Data2')

            with ExcelFile(pth) as reader:
                found_df = reader.parse('Data1')
                found_df2 = reader.parse('Data2')
                tm.assert_frame_equal(found_df, self.frame)
                tm.assert_frame_equal(found_df2, self.frame2)

    def test_roundtrip(self):
        _skip_if_no_xlrd()

        with ensure_clean(self.ext) as path:
            self.frame['A'][:5] = nan

            self.frame.to_excel(path, 'test1')
            self.frame.to_excel(path, 'test1', cols=['A', 'B'])
            self.frame.to_excel(path, 'test1', header=False)
            self.frame.to_excel(path, 'test1', index=False)

            # test roundtrip
            self.frame.to_excel(path, 'test1')
            recons = read_excel(path, 'test1', index_col=0)
            tm.assert_frame_equal(self.frame, recons)

            self.frame.to_excel(path, 'test1', index=False)
            recons = read_excel(path, 'test1', index_col=None)
            recons.index = self.frame.index
            tm.assert_frame_equal(self.frame, recons)

            self.frame.to_excel(path, 'test1', na_rep='NA')
            recons = read_excel(path, 'test1', index_col=0, na_values=['NA'])
            tm.assert_frame_equal(self.frame, recons)

            # GH 3611
            self.frame.to_excel(path, 'test1', na_rep='88')
            recons = read_excel(path, 'test1', index_col=0, na_values=['88'])
            tm.assert_frame_equal(self.frame, recons)

            self.frame.to_excel(path, 'test1', na_rep='88')
            recons = read_excel(path, 'test1', index_col=0,
                                na_values=[88, 88.0])
            tm.assert_frame_equal(self.frame, recons)

    def test_mixed(self):
        _skip_if_no_xlrd()

        with ensure_clean(self.ext) as path:
            self.mixed_frame.to_excel(path, 'test1')
            reader = ExcelFile(path)
            recons = reader.parse('test1', index_col=0)
            tm.assert_frame_equal(self.mixed_frame, recons)

    def test_tsframe(self):
        _skip_if_no_xlrd()

        df = tm.makeTimeDataFrame()[:5]

        with ensure_clean(self.ext) as path:
            df.to_excel(path, 'test1')
            reader = ExcelFile(path)
            recons = reader.parse('test1')
            tm.assert_frame_equal(df, recons)

    def test_basics_with_nan(self):
        _skip_if_no_xlrd()
        with ensure_clean(self.ext) as path:
            self.frame['A'][:5] = nan
            self.frame.to_excel(path, 'test1')
            self.frame.to_excel(path, 'test1', cols=['A', 'B'])
            self.frame.to_excel(path, 'test1', header=False)
            self.frame.to_excel(path, 'test1', index=False)

    def test_int_types(self):
        _skip_if_no_xlrd()

        for np_type in (np.int8, np.int16, np.int32, np.int64):

            with ensure_clean(self.ext) as path:
                # Test np.int values read come back as int (rather than float
                # which is Excel's format).
                frame = DataFrame(np.random.randint(-10, 10, size=(10, 2)),
                                  dtype=np_type)
                frame.to_excel(path, 'test1')
                reader = ExcelFile(path)
                recons = reader.parse('test1')
                int_frame = frame.astype(np.int64)
                tm.assert_frame_equal(int_frame, recons)
                recons2 = read_excel(path, 'test1')
                tm.assert_frame_equal(int_frame, recons2)

                # test with convert_float=False comes back as float
                float_frame = frame.astype(float)
                recons = read_excel(path, 'test1', convert_float=False)
                tm.assert_frame_equal(recons, float_frame)

    def test_float_types(self):
        _skip_if_no_xlrd()

        for np_type in (np.float16, np.float32, np.float64):
            with ensure_clean(self.ext) as path:
                # Test np.float values read come back as float.
                frame = DataFrame(np.random.random_sample(10), dtype=np_type)
                frame.to_excel(path, 'test1')
                reader = ExcelFile(path)
                recons = reader.parse('test1').astype(np_type)
                tm.assert_frame_equal(frame, recons, check_dtype=False)

    def test_bool_types(self):
        _skip_if_no_xlrd()

        for np_type in (np.bool8, np.bool_):
            with ensure_clean(self.ext) as path:
                # Test np.bool values read come back as float.
                frame = (DataFrame([1, 0, True, False], dtype=np_type))
                frame.to_excel(path, 'test1')
                reader = ExcelFile(path)
                recons = reader.parse('test1').astype(np_type)
                tm.assert_frame_equal(frame, recons)

    def test_sheets(self):
        _skip_if_no_xlrd()

        with ensure_clean(self.ext) as path:
            self.frame['A'][:5] = nan

            self.frame.to_excel(path, 'test1')
            self.frame.to_excel(path, 'test1', cols=['A', 'B'])
            self.frame.to_excel(path, 'test1', header=False)
            self.frame.to_excel(path, 'test1', index=False)

            # Test writing to separate sheets
            writer = ExcelWriter(path)
            self.frame.to_excel(writer, 'test1')
            self.tsframe.to_excel(writer, 'test2')
            writer.save()
            reader = ExcelFile(path)
            recons = reader.parse('test1', index_col=0)
            tm.assert_frame_equal(self.frame, recons)
            recons = reader.parse('test2', index_col=0)
            tm.assert_frame_equal(self.tsframe, recons)
            np.testing.assert_equal(2, len(reader.sheet_names))
            np.testing.assert_equal('test1', reader.sheet_names[0])
            np.testing.assert_equal('test2', reader.sheet_names[1])

    def test_colaliases(self):
        _skip_if_no_xlrd()

        with ensure_clean(self.ext) as path:
            self.frame['A'][:5] = nan

            self.frame.to_excel(path, 'test1')
            self.frame.to_excel(path, 'test1', cols=['A', 'B'])
            self.frame.to_excel(path, 'test1', header=False)
            self.frame.to_excel(path, 'test1', index=False)

            # column aliases
            col_aliases = Index(['AA', 'X', 'Y', 'Z'])
            self.frame2.to_excel(path, 'test1', header=col_aliases)
            reader = ExcelFile(path)
            rs = reader.parse('test1', index_col=0)
            xp = self.frame2.copy()
            xp.columns = col_aliases
            tm.assert_frame_equal(xp, rs)

    def test_roundtrip_indexlabels(self):
        _skip_if_no_xlrd()

        with ensure_clean(self.ext) as path:

            self.frame['A'][:5] = nan

            self.frame.to_excel(path, 'test1')
            self.frame.to_excel(path, 'test1', cols=['A', 'B'])
            self.frame.to_excel(path, 'test1', header=False)
            self.frame.to_excel(path, 'test1', index=False)

            # test index_label
            frame = (DataFrame(np.random.randn(10, 2)) >= 0)
            frame.to_excel(path, 'test1',
                           index_label=['test'],
                           merge_cells=self.merge_cells)
            reader = ExcelFile(path)
            recons = reader.parse('test1',
                                  index_col=0,
                                  has_index_names=self.merge_cells
                                  ).astype(np.int64)
            frame.index.names = ['test']
            self.assertEqual(frame.index.names, recons.index.names)

            frame = (DataFrame(np.random.randn(10, 2)) >= 0)
            frame.to_excel(path,
                           'test1',
                           index_label=['test', 'dummy', 'dummy2'],
                           merge_cells=self.merge_cells)
            reader = ExcelFile(path)
            recons = reader.parse('test1',
                                  index_col=0,
                                  has_index_names=self.merge_cells
                                  ).astype(np.int64)
            frame.index.names = ['test']
            self.assertEqual(frame.index.names, recons.index.names)

            frame = (DataFrame(np.random.randn(10, 2)) >= 0)
            frame.to_excel(path,
                           'test1',
                           index_label='test',
                           merge_cells=self.merge_cells)
            reader = ExcelFile(path)
            recons = reader.parse('test1',
                                  index_col=0,
                                  has_index_names=self.merge_cells
                                  ).astype(np.int64)
            frame.index.names = ['test']
            tm.assert_frame_equal(frame, recons.astype(bool))

        with ensure_clean(self.ext) as path:

            self.frame.to_excel(path,
                                'test1',
                                cols=['A', 'B', 'C', 'D'],
                                index=False, merge_cells=self.merge_cells)
            # take 'A' and 'B' as indexes (same row as cols 'C', 'D')
            df = self.frame.copy()
            df = df.set_index(['A', 'B'])

            reader = ExcelFile(path)
            recons = reader.parse('test1', index_col=[0, 1])
            tm.assert_frame_equal(df, recons, check_less_precise=True)

    def test_excel_roundtrip_indexname(self):
        _skip_if_no_xlrd()

        df = DataFrame(np.random.randn(10, 4))
        df.index.name = 'foo'

        with ensure_clean(self.ext) as path:
            df.to_excel(path, merge_cells=self.merge_cells)

            xf = ExcelFile(path)
            result = xf.parse(xf.sheet_names[0],
                              index_col=0,
                              has_index_names=self.merge_cells)

            tm.assert_frame_equal(result, df)
            self.assertEqual(result.index.name, 'foo')

    def test_excel_roundtrip_datetime(self):
        _skip_if_no_xlrd()

        # datetime.date, not sure what to test here exactly
        tsf = self.tsframe.copy()
        with ensure_clean(self.ext) as path:

            tsf.index = [x.date() for x in self.tsframe.index]
            tsf.to_excel(path, 'test1', merge_cells=self.merge_cells)
            reader = ExcelFile(path)
            recons = reader.parse('test1')
            tm.assert_frame_equal(self.tsframe, recons)

    # GH4133 - excel output format strings
    def test_excel_date_datetime_format(self):
        _skip_if_no_xlrd()
        df = DataFrame([[date(2014, 1, 31),
                         date(1999, 9, 24)],
                        [datetime(1998, 5, 26, 23, 33, 4),
                         datetime(2014, 2, 28, 13, 5, 13)]],
                       index=['DATE', 'DATETIME'], columns=['X', 'Y'])
        df_expected = DataFrame([[datetime(2014, 1, 31),
                                  datetime(1999, 9, 24)],
                                 [datetime(1998, 5, 26, 23, 33, 4),
                                  datetime(2014, 2, 28, 13, 5, 13)]],
                                index=['DATE', 'DATETIME'], columns=['X', 'Y'])

        with ensure_clean(self.ext) as filename1:
            with ensure_clean(self.ext) as filename2:
                writer1 = ExcelWriter(filename1)
                writer2 = ExcelWriter(filename2,
                  date_format='DD.MM.YYYY',
                  datetime_format='DD.MM.YYYY HH-MM-SS')

                df.to_excel(writer1, 'test1')
                df.to_excel(writer2, 'test1')

                writer1.close()
                writer2.close()

                reader1 = ExcelFile(filename1)
                reader2 = ExcelFile(filename2)

                rs1 = reader1.parse('test1', index_col=None)
                rs2 = reader2.parse('test1', index_col=None)

                tm.assert_frame_equal(rs1, rs2)

                # since the reader returns a datetime object for dates, we need
                # to use df_expected to check the result
                tm.assert_frame_equal(rs2, df_expected)

    def test_to_excel_periodindex(self):
        _skip_if_no_xlrd()

        frame = self.tsframe
        xp = frame.resample('M', kind='period')

        with ensure_clean(self.ext) as path:
            xp.to_excel(path, 'sht1')

            reader = ExcelFile(path)
            rs = reader.parse('sht1', index_col=0, parse_dates=True)
            tm.assert_frame_equal(xp, rs.to_period('M'))

    def test_to_excel_multiindex(self):
        _skip_if_no_xlrd()

        frame = self.frame
        arrays = np.arange(len(frame.index) * 2).reshape(2, -1)
        new_index = MultiIndex.from_arrays(arrays,
                                           names=['first', 'second'])
        frame.index = new_index

        with ensure_clean(self.ext) as path:
            frame.to_excel(path, 'test1', header=False)
            frame.to_excel(path, 'test1', cols=['A', 'B'])

            # round trip
            frame.to_excel(path, 'test1', merge_cells=self.merge_cells)
            reader = ExcelFile(path)
            df = reader.parse('test1', index_col=[0, 1],
                              parse_dates=False,
                              has_index_names=self.merge_cells)
            tm.assert_frame_equal(frame, df)
            self.assertEqual(frame.index.names, df.index.names)

    def test_to_excel_multiindex_dates(self):
        _skip_if_no_xlrd()

        # try multiindex with dates
        tsframe = self.tsframe.copy()
        new_index = [tsframe.index, np.arange(len(tsframe.index))]
        tsframe.index = MultiIndex.from_arrays(new_index)

        with ensure_clean(self.ext) as path:
            tsframe.index.names = ['time', 'foo']
            tsframe.to_excel(path, 'test1', merge_cells=self.merge_cells)
            reader = ExcelFile(path)
            recons = reader.parse('test1',
                                  index_col=[0, 1],
                                  has_index_names=self.merge_cells)

            tm.assert_frame_equal(tsframe, recons)
            self.assertEquals(recons.index.names, ('time', 'foo'))

    def test_to_excel_multiindex_no_write_index(self):
        _skip_if_no_xlrd()

        # Test writing and re-reading a MI witout the index. GH 5616.

        # Initial non-MI frame.
        frame1 = pd.DataFrame({'a': [10, 20], 'b': [30, 40], 'c': [50, 60]})

        # Add a MI.
        frame2 = frame1.copy()
        multi_index = pd.MultiIndex.from_tuples([(70, 80), (90, 100)])
        frame2.index = multi_index

        with ensure_clean(self.ext) as path:

            # Write out to Excel without the index.
            frame2.to_excel(path, 'test1', index=False)

            # Read it back in.
            reader = ExcelFile(path)
            frame3 = reader.parse('test1')

            # Test that it is the same as the initial frame.
            tm.assert_frame_equal(frame1, frame3)

    def test_to_excel_float_format(self):
        _skip_if_no_xlrd()

        df = DataFrame([[0.123456, 0.234567, 0.567567],
                        [12.32112, 123123.2, 321321.2]],
                        index=['A', 'B'], columns=['X', 'Y', 'Z'])

        with ensure_clean(self.ext) as filename:
            df.to_excel(filename, 'test1', float_format='%.2f')

            reader = ExcelFile(filename)
            rs = reader.parse('test1', index_col=None)
            xp = DataFrame([[0.12, 0.23, 0.57],
                            [12.32, 123123.20, 321321.20]],
                            index=['A', 'B'], columns=['X', 'Y', 'Z'])
            tm.assert_frame_equal(rs, xp)

    def test_to_excel_unicode_filename(self):
        _skip_if_no_xlrd()
        with ensure_clean(u('\u0192u.') + self.ext) as filename:
            try:
                f = open(filename, 'wb')
            except UnicodeEncodeError:
                raise nose.SkipTest('no unicode file names on this system')
            else:
                f.close()

            df = DataFrame([[0.123456, 0.234567, 0.567567],
                            [12.32112, 123123.2, 321321.2]],
                            index=['A', 'B'], columns=['X', 'Y', 'Z'])

            df.to_excel(filename, 'test1', float_format='%.2f')

            reader = ExcelFile(filename)
            rs = reader.parse('test1', index_col=None)
            xp = DataFrame([[0.12, 0.23, 0.57],
                            [12.32, 123123.20, 321321.20]],
                            index=['A', 'B'], columns=['X', 'Y', 'Z'])
            tm.assert_frame_equal(rs, xp)

    # def test_to_excel_header_styling_xls(self):

    #     import StringIO
    #     s = StringIO(
    #     """Date,ticker,type,value
    #     2001-01-01,x,close,12.2
    #     2001-01-01,x,open ,12.1
    #     2001-01-01,y,close,12.2
    #     2001-01-01,y,open ,12.1
    #     2001-02-01,x,close,12.2
    #     2001-02-01,x,open ,12.1
    #     2001-02-01,y,close,12.2
    #     2001-02-01,y,open ,12.1
    #     2001-03-01,x,close,12.2
    #     2001-03-01,x,open ,12.1
    #     2001-03-01,y,close,12.2
    #     2001-03-01,y,open ,12.1""")
    #     df = read_csv(s, parse_dates=["Date"])
    #     pdf = df.pivot_table(values="value", rows=["ticker"],
    #                                          cols=["Date", "type"])

    #     try:
    #         import xlwt
    #         import xlrd
    #     except ImportError:
    #         raise nose.SkipTest

    #     filename = '__tmp_to_excel_header_styling_xls__.xls'
    #     pdf.to_excel(filename, 'test1')

    #     wbk = xlrd.open_workbook(filename,
    #                              formatting_info=True)
    #     self.assertEquals(["test1"], wbk.sheet_names())
    #     ws = wbk.sheet_by_name('test1')
    #     self.assertEquals([(0, 1, 5, 7), (0, 1, 3, 5), (0, 1, 1, 3)],
    #                       ws.merged_cells)
    #     for i in range(0, 2):
    #         for j in range(0, 7):
    #             xfx = ws.cell_xf_index(0, 0)
    #             cell_xf = wbk.xf_list[xfx]
    #             font = wbk.font_list
    #             self.assertEquals(1, font[cell_xf.font_index].bold)
    #             self.assertEquals(1, cell_xf.border.top_line_style)
    #             self.assertEquals(1, cell_xf.border.right_line_style)
    #             self.assertEquals(1, cell_xf.border.bottom_line_style)
    #             self.assertEquals(1, cell_xf.border.left_line_style)
    #             self.assertEquals(2, cell_xf.alignment.hor_align)
    #     os.remove(filename)
    # def test_to_excel_header_styling_xlsx(self):
    #     import StringIO
    #     s = StringIO(
    #     """Date,ticker,type,value
    #     2001-01-01,x,close,12.2
    #     2001-01-01,x,open ,12.1
    #     2001-01-01,y,close,12.2
    #     2001-01-01,y,open ,12.1
    #     2001-02-01,x,close,12.2
    #     2001-02-01,x,open ,12.1
    #     2001-02-01,y,close,12.2
    #     2001-02-01,y,open ,12.1
    #     2001-03-01,x,close,12.2
    #     2001-03-01,x,open ,12.1
    #     2001-03-01,y,close,12.2
    #     2001-03-01,y,open ,12.1""")
    #     df = read_csv(s, parse_dates=["Date"])
    #     pdf = df.pivot_table(values="value", rows=["ticker"],
    #                                          cols=["Date", "type"])
    #     try:
    #         import openpyxl
    #         from openpyxl.cell import get_column_letter
    #     except ImportError:
    #         raise nose.SkipTest
    #     if openpyxl.__version__ < '1.6.1':
    #         raise nose.SkipTest
    #     # test xlsx_styling
    #     filename = '__tmp_to_excel_header_styling_xlsx__.xlsx'
    #     pdf.to_excel(filename, 'test1')
    #     wbk = openpyxl.load_workbook(filename)
    #     self.assertEquals(["test1"], wbk.get_sheet_names())
    #     ws = wbk.get_sheet_by_name('test1')
    #     xlsaddrs = ["%s2" % chr(i) for i in range(ord('A'), ord('H'))]
    #     xlsaddrs += ["A%s" % i for i in range(1, 6)]
    #     xlsaddrs += ["B1", "D1", "F1"]
    #     for xlsaddr in xlsaddrs:
    #         cell = ws.cell(xlsaddr)
    #         self.assertTrue(cell.style.font.bold)
    #         self.assertEquals(openpyxl.style.Border.BORDER_THIN,
    #                           cell.style.borders.top.border_style)
    #         self.assertEquals(openpyxl.style.Border.BORDER_THIN,
    #                           cell.style.borders.right.border_style)
    #         self.assertEquals(openpyxl.style.Border.BORDER_THIN,
    #                           cell.style.borders.bottom.border_style)
    #         self.assertEquals(openpyxl.style.Border.BORDER_THIN,
    #                           cell.style.borders.left.border_style)
    #         self.assertEquals(openpyxl.style.Alignment.HORIZONTAL_CENTER,
    #                           cell.style.alignment.horizontal)
    #     mergedcells_addrs = ["C1", "E1", "G1"]
    #     for maddr in mergedcells_addrs:
    #         self.assertTrue(ws.cell(maddr).merged)
    #     os.remove(filename)

    def test_excel_010_hemstring(self):
        _skip_if_no_xlrd()

        if self.merge_cells:
            raise nose.SkipTest('Skip tests for merged MI format.')

        from pandas.util.testing import makeCustomDataframe as mkdf
        # ensure limited functionality in 0.10
        # override of #2370 until sorted out in 0.11

        def roundtrip(df, header=True, parser_hdr=0):

            with ensure_clean(self.ext) as path:
                df.to_excel(path, header=header, merge_cells=self.merge_cells)
                xf = pd.ExcelFile(path)
                res = xf.parse(xf.sheet_names[0], header=parser_hdr)
                return res

        nrows = 5
        ncols = 3

        for i in range(1, 4):  # row multindex upto nlevel=3
            for j in range(1, 4):  # col ""
                df = mkdf(nrows, ncols, r_idx_nlevels=i, c_idx_nlevels=j)
                res = roundtrip(df)
                # shape
                self.assertEqual(res.shape, (nrows, ncols + i))

                # no nans
                for r in range(len(res.index)):
                    for c in range(len(res.columns)):
                        self.assertTrue(res.ix[r, c] is not np.nan)

        for i in range(1, 4):  # row multindex upto nlevel=3
            for j in range(1, 4):  # col ""
                df = mkdf(nrows, ncols, r_idx_nlevels=i, c_idx_nlevels=j)
                res = roundtrip(df, False)
                # shape
                self.assertEqual(res.shape, (
                    nrows - 1, ncols + i))  # first row taken as columns

                # no nans
                for r in range(len(res.index)):
                    for c in range(len(res.columns)):
                        self.assertTrue(res.ix[r, c] is not np.nan)

        res = roundtrip(DataFrame([0]))
        self.assertEqual(res.shape, (1, 1))
        self.assertTrue(res.ix[0, 0] is not np.nan)

        res = roundtrip(DataFrame([0]), False, None)
        self.assertEqual(res.shape, (1, 2))
        self.assertTrue(res.ix[0, 0] is not np.nan)

    def test_duplicated_columns(self):
        # Test for issue #5235.
        _skip_if_no_xlrd()

        with ensure_clean(self.ext) as path:
            write_frame = DataFrame([[1, 2, 3], [1, 2, 3], [1, 2, 3]])
            colnames = ['A', 'B', 'B']

            write_frame.columns = colnames
            write_frame.to_excel(path, 'test1')

            read_frame = read_excel(path, 'test1')
            read_frame.columns = colnames

            tm.assert_frame_equal(write_frame, read_frame)

    def test_swapped_columns(self):
        # Test for issue #5427.
        _skip_if_no_xlrd()

        with ensure_clean(self.ext) as path:
            write_frame = DataFrame({'A': [1, 1, 1],
                                     'B': [2, 2, 2]})
            write_frame.to_excel(path, 'test1', cols=['B', 'A'])

            read_frame = read_excel(path, 'test1', header=0)

            tm.assert_series_equal(write_frame['A'], read_frame['A'])
            tm.assert_series_equal(write_frame['B'], read_frame['B'])


class OpenpyxlTests(ExcelWriterBase, tm.TestCase):
    ext = '.xlsx'
    engine_name = 'openpyxl'
    check_skip = staticmethod(_skip_if_no_openpyxl)

    def test_to_excel_styleconverter(self):
        _skip_if_no_openpyxl()

        import openpyxl

        hstyle = {"font": {"bold": True},
                  "borders": {"top": "thin",
                              "right": "thin",
                              "bottom": "thin",
                              "left": "thin"},
                  "alignment": {"horizontal": "center", "vertical": "top"}}

        xlsx_style = _OpenpyxlWriter._convert_to_style(hstyle)
        self.assertTrue(xlsx_style.font.bold)
        self.assertEquals(openpyxl.style.Border.BORDER_THIN,
                          xlsx_style.borders.top.border_style)
        self.assertEquals(openpyxl.style.Border.BORDER_THIN,
                          xlsx_style.borders.right.border_style)
        self.assertEquals(openpyxl.style.Border.BORDER_THIN,
                          xlsx_style.borders.bottom.border_style)
        self.assertEquals(openpyxl.style.Border.BORDER_THIN,
                          xlsx_style.borders.left.border_style)
        self.assertEquals(openpyxl.style.Alignment.HORIZONTAL_CENTER,
                          xlsx_style.alignment.horizontal)
        self.assertEquals(openpyxl.style.Alignment.VERTICAL_TOP,
                          xlsx_style.alignment.vertical)


class XlwtTests(ExcelWriterBase, tm.TestCase):
    ext = '.xls'
    engine_name = 'xlwt'
    check_skip = staticmethod(_skip_if_no_xlwt)

    def test_to_excel_styleconverter(self):
        _skip_if_no_xlwt()

        import xlwt

        hstyle = {"font": {"bold": True},
                  "borders": {"top": "thin",
                              "right": "thin",
                              "bottom": "thin",
                              "left": "thin"},
                  "alignment": {"horizontal": "center", "vertical": "top"}}

        xls_style = _XlwtWriter._convert_to_style(hstyle)
        self.assertTrue(xls_style.font.bold)
        self.assertEquals(xlwt.Borders.THIN, xls_style.borders.top)
        self.assertEquals(xlwt.Borders.THIN, xls_style.borders.right)
        self.assertEquals(xlwt.Borders.THIN, xls_style.borders.bottom)
        self.assertEquals(xlwt.Borders.THIN, xls_style.borders.left)
        self.assertEquals(xlwt.Alignment.HORZ_CENTER, xls_style.alignment.horz)
        self.assertEquals(xlwt.Alignment.VERT_TOP, xls_style.alignment.vert)


class XlsxWriterTests(ExcelWriterBase, tm.TestCase):
    ext = '.xlsx'
    engine_name = 'xlsxwriter'
    check_skip = staticmethod(_skip_if_no_xlsxwriter)


class OpenpyxlTests_NoMerge(ExcelWriterBase, tm.TestCase):
    ext = '.xlsx'
    engine_name = 'openpyxl'
    check_skip = staticmethod(_skip_if_no_openpyxl)

    # Test < 0.13 non-merge behaviour for MultiIndex and Hierarchical Rows.
    merge_cells = False


class XlwtTests_NoMerge(ExcelWriterBase, tm.TestCase):
    ext = '.xls'
    engine_name = 'xlwt'
    check_skip = staticmethod(_skip_if_no_xlwt)

    # Test < 0.13 non-merge behaviour for MultiIndex and Hierarchical Rows.
    merge_cells = False


class XlsxWriterTests_NoMerge(ExcelWriterBase, tm.TestCase):
    ext = '.xlsx'
    engine_name = 'xlsxwriter'
    check_skip = staticmethod(_skip_if_no_xlsxwriter)

    # Test < 0.13 non-merge behaviour for MultiIndex and Hierarchical Rows.
    merge_cells = False


class ExcelWriterEngineTests(tm.TestCase):
    def test_ExcelWriter_dispatch(self):
        with tm.assertRaisesRegexp(ValueError, 'No engine'):
            ExcelWriter('nothing')

        try:
            import xlsxwriter
            writer_klass = _XlsxWriter
        except ImportError:
            _skip_if_no_openpyxl()
            writer_klass = _OpenpyxlWriter

        with ensure_clean('.xlsx') as path:
            writer = ExcelWriter(path)
            tm.assert_isinstance(writer, writer_klass)

        _skip_if_no_xlwt()
        with ensure_clean('.xls') as path:
            writer = ExcelWriter(path)
            tm.assert_isinstance(writer, _XlwtWriter)

    def test_register_writer(self):
        # some awkward mocking to test out dispatch and such actually works
        called_save = []
        called_write_cells = []

        class DummyClass(ExcelWriter):
            called_save = False
            called_write_cells = False
            supported_extensions = ['test', 'xlsx', 'xls']
            engine = 'dummy'

            def save(self):
                called_save.append(True)

            def write_cells(self, *args, **kwargs):
                called_write_cells.append(True)

        def check_called(func):
            func()
            self.assert_(len(called_save) >= 1)
            self.assert_(len(called_write_cells) >= 1)
            del called_save[:]
            del called_write_cells[:]

        register_writer(DummyClass)
        writer = ExcelWriter('something.test')
        tm.assert_isinstance(writer, DummyClass)
        df = tm.makeCustomDataframe(1, 1)
        panel = tm.makePanel()
        func = lambda: df.to_excel('something.test')
        check_called(func)
        check_called(lambda: panel.to_excel('something.test'))
        val = get_option('io.excel.xlsx.writer')
        set_option('io.excel.xlsx.writer', 'dummy')
        check_called(lambda: df.to_excel('something.xlsx'))
        check_called(lambda: df.to_excel('something.xls', engine='dummy'))
        set_option('io.excel.xlsx.writer', val)

if __name__ == '__main__':
    nose.runmodule(argv=[__file__, '-vvs', '-x', '--pdb', '--pdb-failure'],
                   exit=False)
