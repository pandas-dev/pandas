"""
Module parse to/from Excel
"""

#----------------------------------------------------------------------
# ExcelFile class
import os
import datetime
import abc
import numpy as np

from pandas.io.parsers import TextParser
from pandas.tseries.period import Period
from pandas import json
from pandas.compat import map, zip, reduce, range, lrange, u, add_metaclass
from pandas.core import config
from pandas.core.common import pprint_thing
import pandas.compat as compat
import pandas.core.common as com
from warnings import warn

__all__ = ["read_excel", "ExcelWriter", "ExcelFile"]

_writer_extensions = ["xlsx", "xls", "xlsm"]
_writers = {}


def register_writer(klass):
    """Adds engine to the excel writer registry. You must use this method to
    integrate with ``to_excel``. Also adds config options for any new
    ``supported_extensions`` defined on the writer."""
    if not compat.callable(klass):
        raise ValueError("Can only register callables as engines")
    engine_name = klass.engine
    _writers[engine_name] = klass
    for ext in klass.supported_extensions:
        if ext.startswith('.'):
            ext = ext[1:]
        if ext not in _writer_extensions:
            config.register_option("io.excel.%s.writer" % ext,
                                   engine_name, validator=str)
            _writer_extensions.append(ext)


def get_writer(engine_name):
    try:
        return _writers[engine_name]
    except KeyError:
        raise ValueError("No Excel writer '%s'" % engine_name)


def read_excel(io, sheetname, **kwds):
    """Read an Excel table into a pandas DataFrame

    Parameters
    ----------
    io : string, file-like object or xlrd workbook
        If a string, expected to be a path to xls or xlsx file
    sheetname : string
         Name of Excel sheet
    header : int, default 0
         Row to use for the column labels of the parsed DataFrame
    skiprows : list-like
        Rows to skip at the beginning (0-indexed)
    skip_footer : int, default 0
        Rows at the end to skip (0-indexed)
    index_col : int, default None
        Column to use as the row labels of the DataFrame. Pass None if
        there is no such column
    parse_cols : int or list, default None
        * If None then parse all columns,
        * If int then indicates last column to be parsed
        * If list of ints then indicates list of column numbers to be parsed
        * If string then indicates comma separated list of column names and
          column ranges (e.g. "A:E" or "A,C,E:F")
    na_values : list-like, default None
        List of additional strings to recognize as NA/NaN
    keep_default_na : bool, default True
        If na_values are specified and keep_default_na is False the default NaN
        values are overridden, otherwise they're appended to
    verbose : boolean, default False
        Indicate number of NA values placed in non-numeric columns
    engine: string, default None
        If io is not a buffer or path, this must be set to identify io.
        Acceptable values are None or xlrd
    convert_float : boolean, default True
        convert integral floats to int (i.e., 1.0 --> 1). If False, all numeric
        data will be read in as floats: Excel stores all numbers as floats
        internally.

    Returns
    -------
    parsed : DataFrame
        DataFrame from the passed in Excel file
    """
    if 'kind' in kwds:
        kwds.pop('kind')
        warn("kind keyword is no longer supported in read_excel and may be "
             "removed in a future version", FutureWarning)

    engine = kwds.pop('engine', None)

    return ExcelFile(io, engine=engine).parse(sheetname=sheetname, **kwds)


class ExcelFile(object):
    """
    Class for parsing tabular excel sheets into DataFrame objects.
    Uses xlrd. See ExcelFile.parse for more documentation

    Parameters
    ----------
    io : string, file-like object or xlrd workbook
        If a string, expected to be a path to xls or xlsx file
    engine: string, default None
        If io is not a buffer or path, this must be set to identify io.
        Acceptable values are None or xlrd
    """
    def __init__(self, io, **kwds):

        import xlrd  # throw an ImportError if we need to

        ver = tuple(map(int, xlrd.__VERSION__.split(".")[:2]))
        if ver < (0, 9):  # pragma: no cover
            raise ImportError("pandas requires xlrd >= 0.9.0 for excel "
                              "support, current version " + xlrd.__VERSION__)

        self.io = io

        engine = kwds.pop('engine', None)

        if engine is not None and engine != 'xlrd':
            raise ValueError("Unknown engine: %s" % engine)

        if isinstance(io, compat.string_types):
            self.book = xlrd.open_workbook(io)
        elif engine == "xlrd" and isinstance(io, xlrd.Book):
            self.book = io
        elif hasattr(io, "read"):
            data = io.read()
            self.book = xlrd.open_workbook(file_contents=data)
        else:
            raise ValueError('Must explicitly set engine if not passing in'
                             ' buffer or path for io.')

    def parse(self, sheetname, header=0, skiprows=None, skip_footer=0,
              index_col=None, parse_cols=None, parse_dates=False,
              date_parser=None, na_values=None, thousands=None, chunksize=None,
              convert_float=True, has_index_names=False, **kwds):
        """Read an Excel table into DataFrame

        Parameters
        ----------
        sheetname : string or integer
            Name of Excel sheet or the page number of the sheet
        header : int, default 0
            Row to use for the column labels of the parsed DataFrame
        skiprows : list-like
            Rows to skip at the beginning (0-indexed)
        skip_footer : int, default 0
            Rows at the end to skip (0-indexed)
        index_col : int, default None
            Column to use as the row labels of the DataFrame. Pass None if
            there is no such column
        parse_cols : int or list, default None
            * If None then parse all columns
            * If int then indicates last column to be parsed
            * If list of ints then indicates list of column numbers to be
              parsed
            * If string then indicates comma separated list of column names and
              column ranges (e.g. "A:E" or "A,C,E:F")
        parse_dates : boolean, default False
            Parse date Excel values,
        date_parser : function default None
            Date parsing function
        na_values : list-like, default None
            List of additional strings to recognize as NA/NaN
        thousands : str, default None
            Thousands separator
        chunksize : int, default None
            Size of file chunk to read for lazy evaluation.
        convert_float : boolean, default True
            convert integral floats to int (i.e., 1.0 --> 1). If False, all
            numeric data will be read in as floats: Excel stores all numbers as
            floats internally.
        has_index_names : boolean, default False
            True if the cols defined in index_col have an index name and are
            not in the header

        Returns
        -------
        parsed : DataFrame
            DataFrame parsed from the Excel file
        """
        skipfooter = kwds.pop('skipfooter', None)
        if skipfooter is not None:
            skip_footer = skipfooter

        return self._parse_excel(sheetname, header=header, skiprows=skiprows,
                                 index_col=index_col,
                                 has_index_names=has_index_names,
                                 parse_cols=parse_cols,
                                 parse_dates=parse_dates,
                                 date_parser=date_parser, na_values=na_values,
                                 thousands=thousands, chunksize=chunksize,
                                 skip_footer=skip_footer,
                                 convert_float=convert_float,
                                 **kwds)

    def _should_parse(self, i, parse_cols):

        def _range2cols(areas):
            """
            Convert comma separated list of column names and column ranges to a
            list of 0-based column indexes.

            >>> _range2cols('A:E')
            [0, 1, 2, 3, 4]
            >>> _range2cols('A,C,Z:AB')
            [0, 2, 25, 26, 27]
            """
            def _excel2num(x):
                "Convert Excel column name like 'AB' to 0-based column index"
                return reduce(lambda s, a: s * 26 + ord(a) - ord('A') + 1,
                              x.upper().strip(), 0) - 1

            cols = []
            for rng in areas.split(','):
                if ':' in rng:
                    rng = rng.split(':')
                    cols += lrange(_excel2num(rng[0]), _excel2num(rng[1]) + 1)
                else:
                    cols.append(_excel2num(rng))
            return cols

        if isinstance(parse_cols, int):
            return i <= parse_cols
        elif isinstance(parse_cols, compat.string_types):
            return i in _range2cols(parse_cols)
        else:
            return i in parse_cols

    def _parse_excel(self, sheetname, header=0, skiprows=None, skip_footer=0,
                     index_col=None, has_index_names=None, parse_cols=None,
                     parse_dates=False, date_parser=None, na_values=None,
                     thousands=None, chunksize=None, convert_float=True,
                     **kwds):
        from xlrd import (xldate_as_tuple, XL_CELL_DATE,
                          XL_CELL_ERROR, XL_CELL_BOOLEAN,
                          XL_CELL_NUMBER)

        datemode = self.book.datemode
        if isinstance(sheetname, compat.string_types):
            sheet = self.book.sheet_by_name(sheetname)
        else:  # assume an integer if not a string
            sheet = self.book.sheet_by_index(sheetname)

        data = []
        should_parse = {}
        for i in range(sheet.nrows):
            row = []
            for j, (value, typ) in enumerate(zip(sheet.row_values(i),
                                                 sheet.row_types(i))):
                if parse_cols is not None and j not in should_parse:
                    should_parse[j] = self._should_parse(j, parse_cols)

                if parse_cols is None or should_parse[j]:
                    if typ == XL_CELL_DATE:
                        dt = xldate_as_tuple(value, datemode)
                        # how to produce this first case?
                        if dt[0] < datetime.MINYEAR:  # pragma: no cover
                            value = datetime.time(*dt[3:])
                        else:
                            value = datetime.datetime(*dt)
                    elif typ == XL_CELL_ERROR:
                        value = np.nan
                    elif typ == XL_CELL_BOOLEAN:
                        value = bool(value)
                    elif convert_float and typ == XL_CELL_NUMBER:
                        # GH5394 - Excel 'numbers' are always floats
                        # it's a minimal perf hit and less suprising
                        val = int(value)
                        if val == value:
                            value = val

                    row.append(value)

            data.append(row)

        if header is not None:
            data[header] = _trim_excel_header(data[header])

        parser = TextParser(data, header=header, index_col=index_col,
                            has_index_names=has_index_names,
                            na_values=na_values,
                            thousands=thousands,
                            parse_dates=parse_dates,
                            date_parser=date_parser,
                            skiprows=skiprows,
                            skip_footer=skip_footer,
                            chunksize=chunksize,
                            **kwds)

        return parser.read()

    @property
    def sheet_names(self):
        return self.book.sheet_names()

    def close(self):
        """close io if necessary"""
        if hasattr(self.io, 'close'):
            self.io.close()

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_value, traceback):
        self.close()


def _trim_excel_header(row):
    # trim header row so auto-index inference works
    # xlrd uses '' , openpyxl None
    while len(row) > 0 and (row[0] == '' or row[0] is None):
        row = row[1:]
    return row


def _conv_value(val):
    # Convert numpy types to Python types for the Excel writers.
    if com.is_integer(val):
        val = int(val)
    elif com.is_float(val):
        val = float(val)
    elif com.is_bool(val):
        val = bool(val)
    elif isinstance(val, Period):
        val = "%s" % val

    return val


@add_metaclass(abc.ABCMeta)
class ExcelWriter(object):
    """
    Class for writing DataFrame objects into excel sheets, default is to use
    xlwt for xls, openpyxl for xlsx.  See DataFrame.to_excel for typical usage.

    Parameters
    ----------
    path : string
        Path to xls or xlsx file.
    engine : string (optional)
        Engine to use for writing. If None, defaults to
        ``io.excel.<extension>.writer``.  NOTE: can only be passed as a keyword
        argument.
    date_format : string, default None
        Format string for dates written into Excel files (e.g. 'YYYY-MM-DD')
    datetime_format : string, default None
        Format string for datetime objects written into Excel files
        (e.g. 'YYYY-MM-DD HH:MM:SS')
    """
    # Defining an ExcelWriter implementation (see abstract methods for more...)

    # - Mandatory
    #   - ``write_cells(self, cells, sheet_name=None, startrow=0, startcol=0)``
    #     --> called to write additional DataFrames to disk
    #   - ``supported_extensions`` (tuple of supported extensions), used to
    #      check that engine supports the given extension.
    #   - ``engine`` - string that gives the engine name. Necessary to
    #     instantiate class directly and bypass ``ExcelWriterMeta`` engine
    #     lookup.
    #   - ``save(self)`` --> called to save file to disk
    # - Mostly mandatory (i.e. should at least exist)
    #   - book, cur_sheet, path

    # - Optional:
    #   - ``__init__(self, path, engine=None, **kwargs)`` --> always called
    #     with path as first argument.

    # You also need to register the class with ``register_writer()``.
    # Technically, ExcelWriter implementations don't need to subclass
    # ExcelWriter.
    def __new__(cls, path, engine=None, **kwargs):
        # only switch class if generic(ExcelWriter)
        if cls == ExcelWriter:
            if engine is None:
                ext = os.path.splitext(path)[-1][1:]
                try:
                    engine = config.get_option('io.excel.%s.writer' % ext)
                except KeyError:
                    error = ValueError("No engine for filetype: '%s'" % ext)
                    raise error
            cls = get_writer(engine)

        return object.__new__(cls)

    # declare external properties you can count on
    book = None
    curr_sheet = None
    path = None

    @abc.abstractproperty
    def supported_extensions(self):
        "extensions that writer engine supports"
        pass

    @abc.abstractproperty
    def engine(self):
        "name of engine"
        pass

    @abc.abstractmethod
    def write_cells(self, cells, sheet_name=None, startrow=0, startcol=0):
        """
        Write given formated cells into Excel an excel sheet

        Parameters
        ----------
        cells : generator
            cell of formated data to save to Excel sheet
        sheet_name : string, default None
            Name of Excel sheet, if None, then use self.cur_sheet
        startrow: upper left cell row to dump data frame
        startcol: upper left cell column to dump data frame
        """
        pass

    @abc.abstractmethod
    def save(self):
        """
        Save workbook to disk.
        """
        pass

    def __init__(self, path, engine=None,
                 date_format=None, datetime_format=None, **engine_kwargs):
        # validate that this engine can handle the extension
        ext = os.path.splitext(path)[-1]
        self.check_extension(ext)

        self.path = path
        self.sheets = {}
        self.cur_sheet = None

        if date_format is None:
            self.date_format = 'YYYY-MM-DD'
        else:
            self.date_format = date_format
        if datetime_format is None:
            self.datetime_format = 'YYYY-MM-DD HH:MM:SS'
        else:
            self.datetime_format = datetime_format

    def _get_sheet_name(self, sheet_name):
        if sheet_name is None:
            sheet_name = self.cur_sheet
        if sheet_name is None:  # pragma: no cover
            raise ValueError('Must pass explicit sheet_name or set '
                             'cur_sheet property')
        return sheet_name

    @classmethod
    def check_extension(cls, ext):
        """checks that path's extension against the Writer's supported
        extensions.  If it isn't supported, raises UnsupportedFiletypeError."""
        if ext.startswith('.'):
            ext = ext[1:]
        if not any(ext in extension for extension in cls.supported_extensions):
            msg = (u("Invalid extension for engine '%s': '%s'") %
                   (pprint_thing(cls.engine), pprint_thing(ext)))
            raise ValueError(msg)
        else:
            return True

    # Allow use as a contextmanager
    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_value, traceback):
        self.close()

    def close(self):
        """synonym for save, to make it more file-like"""
        return self.save()


class _OpenpyxlWriter(ExcelWriter):
    engine = 'openpyxl'
    supported_extensions = ('.xlsx', '.xlsm')

    def __init__(self, path, engine=None, **engine_kwargs):
        # Use the openpyxl module as the Excel writer.
        from openpyxl.workbook import Workbook

        super(_OpenpyxlWriter, self).__init__(path, **engine_kwargs)

        # Create workbook object with default optimized_write=True.
        self.book = Workbook()
        # Openpyxl 1.6.1 adds a dummy sheet. We remove it.
        if self.book.worksheets:
            self.book.remove_sheet(self.book.worksheets[0])

    def save(self):
        """
        Save workbook to disk.
        """
        return self.book.save(self.path)

    def write_cells(self, cells, sheet_name=None, startrow=0, startcol=0):
        # Write the frame cells using openpyxl.
        from openpyxl.cell import get_column_letter

        sheet_name = self._get_sheet_name(sheet_name)

        if sheet_name in self.sheets:
            wks = self.sheets[sheet_name]
        else:
            wks = self.book.create_sheet()
            wks.title = sheet_name
            self.sheets[sheet_name] = wks

        for cell in cells:
            colletter = get_column_letter(startcol + cell.col + 1)
            xcell = wks.cell("%s%s" % (colletter, startrow + cell.row + 1))
            xcell.value = _conv_value(cell.val)
            style = None
            if cell.style:
                style = self._convert_to_style(cell.style)
                for field in style.__fields__:
                    xcell.style.__setattr__(field,
                                            style.__getattribute__(field))

            if isinstance(cell.val, datetime.datetime):
                xcell.style.number_format.format_code = self.datetime_format
            elif isinstance(cell.val, datetime.date):
                xcell.style.number_format.format_code = self.date_format

            if cell.mergestart is not None and cell.mergeend is not None:
                cletterstart = get_column_letter(startcol + cell.col + 1)
                cletterend = get_column_letter(startcol + cell.mergeend + 1)

                wks.merge_cells('%s%s:%s%s' % (cletterstart,
                                               startrow + cell.row + 1,
                                               cletterend,
                                               startrow + cell.mergestart + 1))

                # Excel requires that the format of the first cell in a merged
                # range is repeated in the rest of the merged range.
                if style:
                    first_row = startrow + cell.row + 1
                    last_row = startrow + cell.mergestart + 1
                    first_col = startcol + cell.col + 1
                    last_col = startcol + cell.mergeend + 1

                    for row in range(first_row, last_row + 1):
                        for col in range(first_col, last_col + 1):
                            if row == first_row and col == first_col:
                                # Ignore first cell. It is already handled.
                                continue
                            colletter = get_column_letter(col)
                            xcell = wks.cell("%s%s" % (colletter, row))
                            for field in style.__fields__:
                                xcell.style.__setattr__(
                                    field, style.__getattribute__(field))

    @classmethod
    def _convert_to_style(cls, style_dict):
        """
        converts a style_dict to an openpyxl style object
        Parameters
        ----------
        style_dict: style dictionary to convert
        """

        from openpyxl.style import Style
        xls_style = Style()
        for key, value in style_dict.items():
            for nk, nv in value.items():
                if key == "borders":
                    (xls_style.borders.__getattribute__(nk)
                     .__setattr__('border_style', nv))
                else:
                    xls_style.__getattribute__(key).__setattr__(nk, nv)

        return xls_style

register_writer(_OpenpyxlWriter)


class _XlwtWriter(ExcelWriter):
    engine = 'xlwt'
    supported_extensions = ('.xls',)

    def __init__(self, path, engine=None, **engine_kwargs):
        # Use the xlwt module as the Excel writer.
        import xlwt

        super(_XlwtWriter, self).__init__(path, **engine_kwargs)

        self.book = xlwt.Workbook()
        self.fm_datetime = xlwt.easyxf(num_format_str=self.datetime_format)
        self.fm_date = xlwt.easyxf(num_format_str=self.date_format)

    def save(self):
        """
        Save workbook to disk.
        """
        return self.book.save(self.path)

    def write_cells(self, cells, sheet_name=None, startrow=0, startcol=0):
        # Write the frame cells using xlwt.

        sheet_name = self._get_sheet_name(sheet_name)

        if sheet_name in self.sheets:
            wks = self.sheets[sheet_name]
        else:
            wks = self.book.add_sheet(sheet_name)
            self.sheets[sheet_name] = wks

        style_dict = {}

        for cell in cells:
            val = _conv_value(cell.val)

            num_format_str = None
            if isinstance(cell.val, datetime.datetime):
                num_format_str = self.datetime_format
            if isinstance(cell.val, datetime.date):
                num_format_str = self.date_format

            stylekey = json.dumps(cell.style)
            if num_format_str:
                stylekey += num_format_str

            if stylekey in style_dict:
                style = style_dict[stylekey]
            else:
                style = self._convert_to_style(cell.style, num_format_str)
                style_dict[stylekey] = style

            if cell.mergestart is not None and cell.mergeend is not None:
                wks.write_merge(startrow + cell.row,
                                startrow + cell.mergestart,
                                startcol + cell.col,
                                startcol + cell.mergeend,
                                val, style)
            else:
                wks.write(startrow + cell.row,
                          startcol + cell.col,
                          val, style)

    @classmethod
    def _style_to_xlwt(cls, item, firstlevel=True, field_sep=',',
                       line_sep=';'):
        """helper which recursively generate an xlwt easy style string
        for example:

            hstyle = {"font": {"bold": True},
            "border": {"top": "thin",
                    "right": "thin",
                    "bottom": "thin",
                    "left": "thin"},
            "align": {"horiz": "center"}}
            will be converted to
            font: bold on; \
                    border: top thin, right thin, bottom thin, left thin; \
                    align: horiz center;
        """
        if hasattr(item, 'items'):
            if firstlevel:
                it = ["%s: %s" % (key, cls._style_to_xlwt(value, False))
                      for key, value in item.items()]
                out = "%s " % (line_sep).join(it)
                return out
            else:
                it = ["%s %s" % (key, cls._style_to_xlwt(value, False))
                      for key, value in item.items()]
                out = "%s " % (field_sep).join(it)
                return out
        else:
            item = "%s" % item
            item = item.replace("True", "on")
            item = item.replace("False", "off")
            return item

    @classmethod
    def _convert_to_style(cls, style_dict, num_format_str=None):
        """
        converts a style_dict to an xlwt style object
        Parameters
        ----------
        style_dict: style dictionary to convert
        num_format_str: optional number format string
        """
        import xlwt

        if style_dict:
            xlwt_stylestr = cls._style_to_xlwt(style_dict)
            style = xlwt.easyxf(xlwt_stylestr, field_sep=',', line_sep=';')
        else:
            style = xlwt.XFStyle()
        if num_format_str is not None:
            style.num_format_str = num_format_str

        return style

register_writer(_XlwtWriter)


class _XlsxWriter(ExcelWriter):
    engine = 'xlsxwriter'
    supported_extensions = ('.xlsx',)

    def __init__(self, path, engine=None,
                 date_format=None, datetime_format=None, **engine_kwargs):
        # Use the xlsxwriter module as the Excel writer.
        import xlsxwriter

        super(_XlsxWriter, self).__init__(path, engine=engine,
            date_format=date_format, datetime_format=datetime_format,
            **engine_kwargs)

        self.book = xlsxwriter.Workbook(path, **engine_kwargs)

    def save(self):
        """
        Save workbook to disk.
        """
        return self.book.close()

    def write_cells(self, cells, sheet_name=None, startrow=0, startcol=0):
        # Write the frame cells using xlsxwriter.

        sheet_name = self._get_sheet_name(sheet_name)

        if sheet_name in self.sheets:
            wks = self.sheets[sheet_name]
        else:
            wks = self.book.add_worksheet(sheet_name)
            self.sheets[sheet_name] = wks

        style_dict = {}

        for cell in cells:
            num_format_str = None
            if isinstance(cell.val, datetime.datetime):
                num_format_str = self.datetime_format
            if isinstance(cell.val, datetime.date):
                num_format_str = self.date_format

            stylekey = json.dumps(cell.style)
            if num_format_str:
                stylekey += num_format_str

            if stylekey in style_dict:
                style = style_dict[stylekey]
            else:
                style = self._convert_to_style(cell.style, num_format_str)
                style_dict[stylekey] = style

            if cell.mergestart is not None and cell.mergeend is not None:
                wks.merge_range(startrow + cell.row,
                                startcol + cell.col,
                                startrow + cell.mergestart,
                                startcol + cell.mergeend,
                                cell.val, style)
            else:
                wks.write(startrow + cell.row,
                          startcol + cell.col,
                          cell.val, style)

    def _convert_to_style(self, style_dict, num_format_str=None):
        """
        converts a style_dict to an xlsxwriter format object
        Parameters
        ----------
        style_dict: style dictionary to convert
        num_format_str: optional number format string
        """

        # Create a XlsxWriter format object.
        xl_format = self.book.add_format()
        
        if num_format_str is not None:
            xl_format.set_num_format(num_format_str)

        if style_dict is None:
            return xl_format
        
        # Map the cell font to XlsxWriter font properties.
        if style_dict.get('font'):
            font = style_dict['font']
            if font.get('bold'):
                xl_format.set_bold()

        # Map the alignment to XlsxWriter alignment properties.
        alignment = style_dict.get('alignment')
        if alignment:
            if (alignment.get('horizontal')
                    and alignment['horizontal'] == 'center'):
                xl_format.set_align('center')
            if (alignment.get('vertical')
                    and alignment['vertical'] == 'top'):
                xl_format.set_align('top')

        # Map the cell borders to XlsxWriter border properties.
        if style_dict.get('borders'):
            xl_format.set_border()

        return xl_format

register_writer(_XlsxWriter)
