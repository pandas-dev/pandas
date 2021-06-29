from pathlib import Path
import re

import numpy as np
import pytest

import pandas as pd
from pandas import DataFrame
import pandas._testing as tm

from pandas.io.excel import (
    ExcelWriter,
    _OpenpyxlWriter,
)

openpyxl = pytest.importorskip("openpyxl")

pytestmark = pytest.mark.parametrize("ext", [".xlsx"])


def test_to_excel_styleconverter(ext):
    from openpyxl import styles

    hstyle = {
        "font": {"color": "00FF0000", "bold": True},
        "borders": {"top": "thin", "right": "thin", "bottom": "thin", "left": "thin"},
        "alignment": {"horizontal": "center", "vertical": "top"},
        "fill": {"patternType": "solid", "fgColor": {"rgb": "006666FF", "tint": 0.3}},
        "number_format": {"format_code": "0.00"},
        "protection": {"locked": True, "hidden": False},
    }

    font_color = styles.Color("00FF0000")
    font = styles.Font(bold=True, color=font_color)
    side = styles.Side(style=styles.borders.BORDER_THIN)
    border = styles.Border(top=side, right=side, bottom=side, left=side)
    alignment = styles.Alignment(horizontal="center", vertical="top")
    fill_color = styles.Color(rgb="006666FF", tint=0.3)
    fill = styles.PatternFill(patternType="solid", fgColor=fill_color)

    number_format = "0.00"

    protection = styles.Protection(locked=True, hidden=False)

    kw = _OpenpyxlWriter._convert_to_style_kwargs(hstyle)
    assert kw["font"] == font
    assert kw["border"] == border
    assert kw["alignment"] == alignment
    assert kw["fill"] == fill
    assert kw["number_format"] == number_format
    assert kw["protection"] == protection


def test_write_cells_merge_styled(ext):
    from pandas.io.formats.excel import ExcelCell

    sheet_name = "merge_styled"

    sty_b1 = {"font": {"color": "00FF0000"}}
    sty_a2 = {"font": {"color": "0000FF00"}}

    initial_cells = [
        ExcelCell(col=1, row=0, val=42, style=sty_b1),
        ExcelCell(col=0, row=1, val=99, style=sty_a2),
    ]

    sty_merged = {"font": {"color": "000000FF", "bold": True}}
    sty_kwargs = _OpenpyxlWriter._convert_to_style_kwargs(sty_merged)
    openpyxl_sty_merged = sty_kwargs["font"]
    merge_cells = [
        ExcelCell(
            col=0, row=0, val="pandas", mergestart=1, mergeend=1, style=sty_merged
        )
    ]

    with tm.ensure_clean(ext) as path:
        with _OpenpyxlWriter(path) as writer:
            writer.write_cells(initial_cells, sheet_name=sheet_name)
            writer.write_cells(merge_cells, sheet_name=sheet_name)

            wks = writer.sheets[sheet_name]
        xcell_b1 = wks["B1"]
        xcell_a2 = wks["A2"]
        assert xcell_b1.font == openpyxl_sty_merged
        assert xcell_a2.font == openpyxl_sty_merged


@pytest.mark.parametrize("write_only", [True, False])
def test_kwargs(ext, write_only):
    # GH 42286
    # openpyxl doesn't utilize kwargs, only test that supplying a kwarg works
    kwargs = {"write_only": write_only}
    with tm.ensure_clean(ext) as f:
        msg = re.escape("Use of **kwargs is deprecated")
        with tm.assert_produces_warning(FutureWarning, match=msg):
            with ExcelWriter(f, engine="openpyxl", **kwargs) as writer:
                # ExcelWriter won't allow us to close without writing something
                DataFrame().to_excel(writer)


@pytest.mark.parametrize("write_only", [True, False])
def test_engine_kwargs(ext, write_only):
    # GH 42286
    # openpyxl doesn't utilize kwargs, only test that supplying a engine_kwarg works
    engine_kwargs = {"write_only": write_only}
    with tm.ensure_clean(ext) as f:
        with ExcelWriter(f, engine="openpyxl", engine_kwargs=engine_kwargs) as writer:
            # ExcelWriter won't allow us to close without writing something
            DataFrame().to_excel(writer)


@pytest.mark.parametrize(
    "mode,expected", [("w", ["baz"]), ("a", ["foo", "bar", "baz"])]
)
def test_write_append_mode(ext, mode, expected):
    df = DataFrame([1], columns=["baz"])

    with tm.ensure_clean(ext) as f:
        wb = openpyxl.Workbook()
        wb.worksheets[0].title = "foo"
        wb.worksheets[0]["A1"].value = "foo"
        wb.create_sheet("bar")
        wb.worksheets[1]["A1"].value = "bar"
        wb.save(f)

        with ExcelWriter(f, engine="openpyxl", mode=mode) as writer:
            df.to_excel(writer, sheet_name="baz", index=False)

        wb2 = openpyxl.load_workbook(f)
        result = [sheet.title for sheet in wb2.worksheets]
        assert result == expected

        for index, cell_value in enumerate(expected):
            assert wb2.worksheets[index]["A1"].value == cell_value


@pytest.mark.parametrize(
    "if_sheet_exists,num_sheets,expected",
    [
        ("new", 2, ["apple", "banana"]),
        ("replace", 1, ["pear"]),
    ],
)
def test_if_sheet_exists_append_modes(ext, if_sheet_exists, num_sheets, expected):
    # GH 40230
    df1 = DataFrame({"fruit": ["apple", "banana"]})
    df2 = DataFrame({"fruit": ["pear"]})

    with tm.ensure_clean(ext) as f:
        df1.to_excel(f, engine="openpyxl", sheet_name="foo", index=False)
        with ExcelWriter(
            f, engine="openpyxl", mode="a", if_sheet_exists=if_sheet_exists
        ) as writer:
            df2.to_excel(writer, sheet_name="foo", index=False)

        wb = openpyxl.load_workbook(f)
        assert len(wb.sheetnames) == num_sheets
        assert wb.sheetnames[0] == "foo"
        result = pd.read_excel(wb, "foo", engine="openpyxl")
        assert list(result["fruit"]) == expected
        if len(wb.sheetnames) == 2:
            result = pd.read_excel(wb, wb.sheetnames[1], engine="openpyxl")
            tm.assert_frame_equal(result, df2)
        wb.close()


@pytest.mark.parametrize(
    "if_sheet_exists,msg",
    [
        (
            "invalid",
            "'invalid' is not valid for if_sheet_exists. Valid options "
            "are 'error', 'new' and 'replace'.",
        ),
        (
            "error",
            "Sheet 'foo' already exists and if_sheet_exists is set to 'error'.",
        ),
        (
            None,
            "Sheet 'foo' already exists and if_sheet_exists is set to 'error'.",
        ),
    ],
)
def test_if_sheet_exists_raises(ext, if_sheet_exists, msg):
    # GH 40230
    df = DataFrame({"fruit": ["pear"]})
    with tm.ensure_clean(ext) as f:
        with pytest.raises(ValueError, match=re.escape(msg)):
            df.to_excel(f, "foo", engine="openpyxl")
            with ExcelWriter(
                f, engine="openpyxl", mode="a", if_sheet_exists=if_sheet_exists
            ) as writer:
                df.to_excel(writer, sheet_name="foo")


def test_to_excel_with_openpyxl_engine(ext):
    # GH 29854
    with tm.ensure_clean(ext) as filename:

        df1 = DataFrame({"A": np.linspace(1, 10, 10)})
        df2 = DataFrame({"B": np.linspace(1, 20, 10)})
        df = pd.concat([df1, df2], axis=1)
        styled = df.style.applymap(
            lambda val: "color: %s" % ("red" if val < 0 else "black")
        ).highlight_max()

        styled.to_excel(filename, engine="openpyxl")


@pytest.mark.parametrize("read_only", [True, False])
def test_read_workbook(datapath, ext, read_only):
    # GH 39528
    filename = datapath("io", "data", "excel", "test1" + ext)
    wb = openpyxl.load_workbook(filename, read_only=read_only)
    result = pd.read_excel(wb, engine="openpyxl")
    wb.close()
    expected = pd.read_excel(filename)
    tm.assert_frame_equal(result, expected)


@pytest.mark.parametrize(
    "header, expected_data",
    [
        (
            0,
            {
                "Title": [np.nan, "A", 1, 2, 3],
                "Unnamed: 1": [np.nan, "B", 4, 5, 6],
                "Unnamed: 2": [np.nan, "C", 7, 8, 9],
            },
        ),
        (2, {"A": [1, 2, 3], "B": [4, 5, 6], "C": [7, 8, 9]}),
    ],
)
@pytest.mark.parametrize(
    "filename", ["dimension_missing", "dimension_small", "dimension_large"]
)
# When read_only is None, use read_excel instead of a workbook
@pytest.mark.parametrize("read_only", [True, False, None])
def test_read_with_bad_dimension(
    datapath, ext, header, expected_data, filename, read_only, request
):
    # GH 38956, 39001 - no/incorrect dimension information
    path = datapath("io", "data", "excel", f"{filename}{ext}")
    if read_only is None:
        result = pd.read_excel(path, header=header)
    else:
        wb = openpyxl.load_workbook(path, read_only=read_only)
        result = pd.read_excel(wb, engine="openpyxl", header=header)
        wb.close()
    expected = DataFrame(expected_data)
    tm.assert_frame_equal(result, expected)


def test_append_mode_file(ext):
    # GH 39576
    df = DataFrame()

    with tm.ensure_clean(ext) as f:
        df.to_excel(f, engine="openpyxl")

        with ExcelWriter(
            f, mode="a", engine="openpyxl", if_sheet_exists="new"
        ) as writer:
            df.to_excel(writer)

        # make sure that zip files are not concatenated by making sure that
        # "docProps/app.xml" only occurs twice in the file
        data = Path(f).read_bytes()
        first = data.find(b"docProps/app.xml")
        second = data.find(b"docProps/app.xml", first + 1)
        third = data.find(b"docProps/app.xml", second + 1)
        assert second != -1 and third == -1


# When read_only is None, use read_excel instead of a workbook
@pytest.mark.parametrize("read_only", [True, False, None])
def test_read_with_empty_trailing_rows(datapath, ext, read_only, request):
    # GH 39181
    path = datapath("io", "data", "excel", f"empty_trailing_rows{ext}")
    if read_only is None:
        result = pd.read_excel(path)
    else:
        wb = openpyxl.load_workbook(path, read_only=read_only)
        result = pd.read_excel(wb, engine="openpyxl")
        wb.close()
    expected = DataFrame(
        {
            "Title": [np.nan, "A", 1, 2, 3],
            "Unnamed: 1": [np.nan, "B", 4, 5, 6],
            "Unnamed: 2": [np.nan, "C", 7, 8, 9],
        }
    )
    tm.assert_frame_equal(result, expected)


# When read_only is None, use read_excel instead of a workbook
@pytest.mark.parametrize("read_only", [True, False, None])
def test_read_empty_with_blank_row(datapath, ext, read_only):
    # GH 39547 - empty excel file with a row that has no data
    path = datapath("io", "data", "excel", f"empty_with_blank_row{ext}")
    if read_only is None:
        result = pd.read_excel(path)
    else:
        wb = openpyxl.load_workbook(path, read_only=read_only)
        result = pd.read_excel(wb, engine="openpyxl")
        wb.close()
    expected = DataFrame()
    tm.assert_frame_equal(result, expected)
