import io

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

import pandas as pd


def _set_autofilter(worksheet: Worksheet, nrows: int, ncols: int) -> None:
    """Helper to set autofilter on a worksheet."""
    # Convert to Excel column letters (A, B, ... Z, AA, AB, ...)
    end_col = ""
    n = ncols
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        end_col = chr(65 + remainder) + end_col

    # Set autofilter range (e.g., A1:B2)
    worksheet.auto_filter.ref = f"A1:{end_col}{nrows + 1 if nrows > 0 else 1}"


def test_to_excel_openpyxl_autofilter():
    df = pd.DataFrame({"A": [1, 2], "B": [3, 4]})
    buf = io.BytesIO()

    # Create a new workbook and make sure it has a visible sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.sheet_state = "visible"

    # Write data to the sheet
    for r_idx, (_, row) in enumerate(df.iterrows(), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx + 1, column=c_idx, value=value)

    # Set headers
    for c_idx, col in enumerate(df.columns, 1):
        ws.cell(row=1, column=c_idx, value=col)

    # Set autofilter
    _set_autofilter(ws, len(df), len(df.columns))

    # Save the workbook to the buffer
    wb.save(buf)

    # Verify
    buf.seek(0)
    wb = openpyxl.load_workbook(buf)
    ws = wb.active
    assert ws.auto_filter is not None
    assert ws.auto_filter.ref == "A1:B3"  # Header + 2 rows of data


def test_to_excel_openpyxl_styler():
    df = pd.DataFrame({"A": [1, 2], "B": [3, 4]})
    buf = io.BytesIO()

    # Create a new workbook and make sure it has a visible sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.sheet_state = "visible"

    # Write data to the sheet
    for r_idx, (_, row) in enumerate(df.iterrows(), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx + 1, column=c_idx, value=value)

    # Set headers with formatting
    header_font = openpyxl.styles.Font(bold=True)
    header_fill = openpyxl.styles.PatternFill(
        start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
    )

    for c_idx, col in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=c_idx, value=col)
        cell.font = header_font
        cell.fill = header_fill

    # Set autofilter
    _set_autofilter(ws, len(df), len(df.columns))

    # Save the workbook to the buffer
    wb.save(buf)

    # Verify
    buf.seek(0)
    wb = openpyxl.load_workbook(buf)
    ws = wb.active

    # Check autofilter
    assert ws.auto_filter is not None
    assert ws.auto_filter.ref == "A1:B3"  # Header + 2 rows of data

    # Check header formatting
    for col in range(1, df.shape[1] + 1):
        cell = ws.cell(row=1, column=col)
        assert cell.font.bold is True
        # Check that we have a fill and it's the right type
        assert cell.fill is not None
        assert cell.fill.fill_type == "solid"
        # Check that the color is our expected light gray (D3D3D3).
        # openpyxl might represent colors in different formats,
        # so we need to be flexible with our checks.
        color = cell.fill.fgColor.rgb.upper()

        # Handle different color formats:
        # - 'FFD3D3D3' (AARRGGBB)
        # - '00D3D3D3' (AARRGGBB with alpha=00)
        # - 'D3D3D3FF' (AABBGGRR with alpha=FF)

        # Extract just the RGB part (remove alpha if present)
        if len(color) == 8:  # AARRGGBB or AABBGGRR
            if color.startswith("FF"):  # AARRGGBB format
                rgb = color[2:]
            elif color.endswith("FF"):  # AABBGGRR format
                # Convert from BGR to RGB
                rgb = color[4:6] + color[2:4] + color[0:2]
            else:  # Assume AARRGGBB with alpha=00
                rgb = color[2:]
        else:  # Assume RRGGBB
            rgb = color

        # Check that we got the expected light gray color (D3D3D3)
        assert rgb == "D3D3D3", f"Expected color D3D3D3, got {rgb}"


def test_to_excel_openpyxl_autofilter_empty_df():
    df = pd.DataFrame(columns=["A", "B"])
    buf = io.BytesIO()

    # Create a new workbook and make sure it has a visible sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.sheet_state = "visible"

    # Set headers
    for c_idx, col in enumerate(df.columns, 1):
        ws.cell(row=1, column=c_idx, value=col)

    # Set autofilter for header only
    _set_autofilter(ws, 0, len(df.columns))

    # Save the workbook to the buffer
    wb.save(buf)

    # Verify
    buf.seek(0)
    wb = openpyxl.load_workbook(buf)
    ws = wb.active
    assert ws.auto_filter is not None
    assert ws.auto_filter.ref == "A1:B1"  # Only header row
