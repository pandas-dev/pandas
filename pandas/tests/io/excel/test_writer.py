from datetime import date, datetime, timedelta
from distutils.version import LooseVersion
from functools import partial
import os
import warnings
from warnings import catch_warnings

import numpy as np
from numpy import nan
import pytest

from pandas.compat import PY36, BytesIO, map, range, u
import pandas.util._test_decorators as td

import pandas as pd
from pandas import DataFrame, Index, MultiIndex
from pandas.core.config import get_option, set_option
import pandas.util.testing as tm
from pandas.util.testing import ensure_clean, makeCustomDataframe as mkdf

from pandas.io.excel import (
    ExcelFile, ExcelWriter, _OpenpyxlWriter, _XlsxWriter, _XlwtWriter,
    read_excel, register_writer)
from pandas.io.formats.excel import ExcelFormatter
from pandas.io.parsers import read_csv

_seriesd = tm.getSeriesData()
_tsd = tm.getTimeSeriesData()
_frame = DataFrame(_seriesd)[:10]
_frame2 = DataFrame(_seriesd, columns=['D', 'C', 'B', 'A'])[:10]
_tsframe = tm.makeTimeDataFrame()[:5]
_mixed_frame = _frame.copy()
_mixed_frame['foo'] = 'bar'


@td.skip_if_no('xlrd', '1.0.0')
class SharedItems(object):

    @pytest.fixture(autouse=True)
    def setup_method(self, datapath):
        self.dirpath = datapath("io", "data")
        self.frame = _frame.copy()
        self.frame2 = _frame2.copy()
        self.tsframe = _tsframe.copy()
        self.mixed_frame = _mixed_frame.copy()

    def get_csv_refdf(self, basename):
        """
        Obtain the reference data from read_csv with the Python engine.
        Parameters
        ----------
        basename : str
            File base name, excluding file extension.
        Returns
        -------
        dfref : DataFrame
        """
        pref = os.path.join(self.dirpath, basename + '.csv')
        dfref = read_csv(pref, index_col=0, parse_dates=True, engine='python')
        return dfref

    def get_excelfile(self, basename, ext):
        """
        Return test data ExcelFile instance.
        Parameters
        ----------
        basename : str
            File base name, excluding file extension.
        Returns
        -------
        excel : io.excel.ExcelFile
        """
        return ExcelFile(os.path.join(self.dirpath, basename + ext))

    def get_exceldf(self, basename, ext, *args, **kwds):
        """
        Return test data DataFrame.
        Parameters
        ----------
        basename : str
            File base name, excluding file extension.
        Returns
        -------
        df : DataFrame
        """
        pth = os.path.join(self.dirpath, basename + ext)
        return read_excel(pth, *args, **kwds)


class _WriterBase(SharedItems):

    @pytest.fixture(autouse=True)
    def set_engine_and_path(self, request, merge_cells, engine, ext):
        """Fixture to set engine and open file for use in each test case
        Rather than requiring `engine=...` to be provided explicitly as an
        argument in each test, this fixture sets a global option to dictate
        which engine should be used to write Excel files. After executing
        the test it rolls back said change to the global option.
        It also uses a context manager to open a temporary excel file for
        the function to write to, accessible via `self.path`
        Notes
        -----
        This fixture will run as part of each test method defined in the
        class and any subclasses, on account of the `autouse=True`
        argument
        """
        option_name = 'io.excel.{ext}.writer'.format(ext=ext.strip('.'))
        prev_engine = get_option(option_name)
        set_option(option_name, engine)
        with ensure_clean(ext) as path:
            self.path = path
            yield
        set_option(option_name, prev_engine)  # Roll back option change


@pytest.mark.parametrize("merge_cells", [True, False])
@pytest.mark.parametrize("engine,ext", [
    pytest.param('openpyxl', '.xlsx', marks=pytest.mark.skipif(
        not td.safe_import('openpyxl'), reason='No openpyxl')),
    pytest.param('openpyxl', '.xlsm', marks=pytest.mark.skipif(
        not td.safe_import('openpyxl'), reason='No openpyxl')),
    pytest.param('xlwt', '.xls', marks=pytest.mark.skipif(
        not td.safe_import('xlwt'), reason='No xlwt')),
    pytest.param('xlsxwriter', '.xlsx', marks=pytest.mark.skipif(
        not td.safe_import('xlsxwriter'), reason='No xlsxwriter'))
])
class TestExcelWriter(_WriterBase):
    # Base class for test cases to run with different Excel writers.

    def test_excel_sheet_by_name_raise(self, *_):
        import xlrd

        gt = DataFrame(np.random.randn(10, 2))
        gt.to_excel(self.path)

        xl = ExcelFile(self.path)
        df = read_excel(xl, 0, index_col=0)

        tm.assert_frame_equal(gt, df)

        with pytest.raises(xlrd.XLRDError):
            read_excel(xl, "0")

    def test_excel_writer_context_manager(self, *_):
        with ExcelWriter(self.path) as writer:
            self.frame.to_excel(writer, "Data1")
            self.frame2.to_excel(writer, "Data2")

        with ExcelFile(self.path) as reader:
            found_df = read_excel(reader, "Data1", index_col=0)
            found_df2 = read_excel(reader, "Data2", index_col=0)

            tm.assert_frame_equal(found_df, self.frame)
            tm.assert_frame_equal(found_df2, self.frame2)

    def test_roundtrip(self, merge_cells, engine, ext):
        self.frame['A'][:5] = nan

        self.frame.to_excel(self.path, 'test1')
        self.frame.to_excel(self.path, 'test1', columns=['A', 'B'])
        self.frame.to_excel(self.path, 'test1', header=False)
        self.frame.to_excel(self.path, 'test1', index=False)

        # test roundtrip
        self.frame.to_excel(self.path, 'test1')
        recons = read_excel(self.path, 'test1', index_col=0)
        tm.assert_frame_equal(self.frame, recons)

        self.frame.to_excel(self.path, 'test1', index=False)
        recons = read_excel(self.path, 'test1', index_col=None)
        recons.index = self.frame.index
        tm.assert_frame_equal(self.frame, recons)

        self.frame.to_excel(self.path, 'test1', na_rep='NA')
        recons = read_excel(self.path, 'test1', index_col=0, na_values=['NA'])
        tm.assert_frame_equal(self.frame, recons)

        # GH 3611
        self.frame.to_excel(self.path, 'test1', na_rep='88')
        recons = read_excel(self.path, 'test1', index_col=0, na_values=['88'])
        tm.assert_frame_equal(self.frame, recons)

        self.frame.to_excel(self.path, 'test1', na_rep='88')
        recons = read_excel(self.path, 'test1', index_col=0,
                            na_values=[88, 88.0])
        tm.assert_frame_equal(self.frame, recons)

        # GH 6573
        self.frame.to_excel(self.path, 'Sheet1')
        recons = read_excel(self.path, index_col=0)
        tm.assert_frame_equal(self.frame, recons)

        self.frame.to_excel(self.path, '0')
        recons = read_excel(self.path, index_col=0)
        tm.assert_frame_equal(self.frame, recons)

        # GH 8825 Pandas Series should provide to_excel method
        s = self.frame["A"]
        s.to_excel(self.path)
        recons = read_excel(self.path, index_col=0)
        tm.assert_frame_equal(s.to_frame(), recons)

    def test_mixed(self, merge_cells, engine, ext):
        self.mixed_frame.to_excel(self.path, 'test1')
        reader = ExcelFile(self.path)
        recons = read_excel(reader, 'test1', index_col=0)
        tm.assert_frame_equal(self.mixed_frame, recons)

    def test_ts_frame(self, *_):
        df = tm.makeTimeDataFrame()[:5]

        df.to_excel(self.path, "test1")
        reader = ExcelFile(self.path)

        recons = read_excel(reader, "test1", index_col=0)
        tm.assert_frame_equal(df, recons)

    def test_basics_with_nan(self, merge_cells, engine, ext):
        self.frame['A'][:5] = nan
        self.frame.to_excel(self.path, 'test1')
        self.frame.to_excel(self.path, 'test1', columns=['A', 'B'])
        self.frame.to_excel(self.path, 'test1', header=False)
        self.frame.to_excel(self.path, 'test1', index=False)

    @pytest.mark.parametrize("np_type", [
        np.int8, np.int16, np.int32, np.int64])
    def test_int_types(self, merge_cells, engine, ext, np_type):
        # Test np.int values read come back as int
        # (rather than float which is Excel's format).
        frame = DataFrame(np.random.randint(-10, 10, size=(10, 2)),
                          dtype=np_type)
        frame.to_excel(self.path, "test1")

        reader = ExcelFile(self.path)
        recons = read_excel(reader, "test1", index_col=0)

        int_frame = frame.astype(np.int64)
        tm.assert_frame_equal(int_frame, recons)

        recons2 = read_excel(self.path, "test1", index_col=0)
        tm.assert_frame_equal(int_frame, recons2)

        # Test with convert_float=False comes back as float.
        float_frame = frame.astype(float)
        recons = read_excel(self.path, "test1",
                            convert_float=False, index_col=0)
        tm.assert_frame_equal(recons, float_frame,
                              check_index_type=False,
                              check_column_type=False)

    @pytest.mark.parametrize("np_type", [
        np.float16, np.float32, np.float64])
    def test_float_types(self, merge_cells, engine, ext, np_type):
        # Test np.float values read come back as float.
        frame = DataFrame(np.random.random_sample(10), dtype=np_type)
        frame.to_excel(self.path, "test1")

        reader = ExcelFile(self.path)
        recons = read_excel(reader, "test1", index_col=0).astype(np_type)

        tm.assert_frame_equal(frame, recons, check_dtype=False)

    @pytest.mark.parametrize("np_type", [np.bool8, np.bool_])
    def test_bool_types(self, merge_cells, engine, ext, np_type):
        # Test np.bool values read come back as float.
        frame = (DataFrame([1, 0, True, False], dtype=np_type))
        frame.to_excel(self.path, "test1")

        reader = ExcelFile(self.path)
        recons = read_excel(reader, "test1", index_col=0).astype(np_type)

        tm.assert_frame_equal(frame, recons)

    def test_inf_roundtrip(self, *_):
        frame = DataFrame([(1, np.inf), (2, 3), (5, -np.inf)])
        frame.to_excel(self.path, "test1")

        reader = ExcelFile(self.path)
        recons = read_excel(reader, "test1", index_col=0)

        tm.assert_frame_equal(frame, recons)

    def test_sheets(self, merge_cells, engine, ext):
        self.frame['A'][:5] = nan

        self.frame.to_excel(self.path, 'test1')
        self.frame.to_excel(self.path, 'test1', columns=['A', 'B'])
        self.frame.to_excel(self.path, 'test1', header=False)
        self.frame.to_excel(self.path, 'test1', index=False)

        # Test writing to separate sheets
        writer = ExcelWriter(self.path)
        self.frame.to_excel(writer, 'test1')
        self.tsframe.to_excel(writer, 'test2')
        writer.save()
        reader = ExcelFile(self.path)
        recons = read_excel(reader, 'test1', index_col=0)
        tm.assert_frame_equal(self.frame, recons)
        recons = read_excel(reader, 'test2', index_col=0)
        tm.assert_frame_equal(self.tsframe, recons)
        assert 2 == len(reader.sheet_names)
        assert 'test1' == reader.sheet_names[0]
        assert 'test2' == reader.sheet_names[1]

    def test_colaliases(self, merge_cells, engine, ext):
        self.frame['A'][:5] = nan

        self.frame.to_excel(self.path, 'test1')
        self.frame.to_excel(self.path, 'test1', columns=['A', 'B'])
        self.frame.to_excel(self.path, 'test1', header=False)
        self.frame.to_excel(self.path, 'test1', index=False)

        # column aliases
        col_aliases = Index(['AA', 'X', 'Y', 'Z'])
        self.frame2.to_excel(self.path, 'test1', header=col_aliases)
        reader = ExcelFile(self.path)
        rs = read_excel(reader, 'test1', index_col=0)
        xp = self.frame2.copy()
        xp.columns = col_aliases
        tm.assert_frame_equal(xp, rs)

    def test_roundtrip_indexlabels(self, merge_cells, engine, ext):
        self.frame['A'][:5] = nan

        self.frame.to_excel(self.path, 'test1')
        self.frame.to_excel(self.path, 'test1', columns=['A', 'B'])
        self.frame.to_excel(self.path, 'test1', header=False)
        self.frame.to_excel(self.path, 'test1', index=False)

        # test index_label
        frame = (DataFrame(np.random.randn(10, 2)) >= 0)
        frame.to_excel(self.path, 'test1',
                       index_label=['test'],
                       merge_cells=merge_cells)
        reader = ExcelFile(self.path)
        recons = read_excel(reader, 'test1',
                            index_col=0,
                            ).astype(np.int64)
        frame.index.names = ['test']
        assert frame.index.names == recons.index.names

        frame = (DataFrame(np.random.randn(10, 2)) >= 0)
        frame.to_excel(self.path,
                       'test1',
                       index_label=['test', 'dummy', 'dummy2'],
                       merge_cells=merge_cells)
        reader = ExcelFile(self.path)
        recons = read_excel(reader, 'test1',
                            index_col=0,
                            ).astype(np.int64)
        frame.index.names = ['test']
        assert frame.index.names == recons.index.names

        frame = (DataFrame(np.random.randn(10, 2)) >= 0)
        frame.to_excel(self.path,
                       'test1',
                       index_label='test',
                       merge_cells=merge_cells)
        reader = ExcelFile(self.path)
        recons = read_excel(reader, 'test1',
                            index_col=0,
                            ).astype(np.int64)
        frame.index.names = ['test']
        tm.assert_frame_equal(frame, recons.astype(bool))

        self.frame.to_excel(self.path,
                            'test1',
                            columns=['A', 'B', 'C', 'D'],
                            index=False, merge_cells=merge_cells)
        # take 'A' and 'B' as indexes (same row as cols 'C', 'D')
        df = self.frame.copy()
        df = df.set_index(['A', 'B'])

        reader = ExcelFile(self.path)
        recons = read_excel(reader, 'test1', index_col=[0, 1])
        tm.assert_frame_equal(df, recons, check_less_precise=True)

    def test_excel_roundtrip_indexname(self, merge_cells, engine, ext):
        df = DataFrame(np.random.randn(10, 4))
        df.index.name = 'foo'

        df.to_excel(self.path, merge_cells=merge_cells)

        xf = ExcelFile(self.path)
        result = read_excel(xf, xf.sheet_names[0],
                            index_col=0)

        tm.assert_frame_equal(result, df)
        assert result.index.name == 'foo'

    def test_excel_roundtrip_datetime(self, merge_cells, *_):
        # datetime.date, not sure what to test here exactly
        tsf = self.tsframe.copy()

        tsf.index = [x.date() for x in self.tsframe.index]
        tsf.to_excel(self.path, "test1", merge_cells=merge_cells)

        reader = ExcelFile(self.path)
        recons = read_excel(reader, "test1", index_col=0)

        tm.assert_frame_equal(self.tsframe, recons)

    def test_excel_date_datetime_format(self, merge_cells, engine, ext):
        # see gh-4133
        #
        # Excel output format strings
        df = DataFrame([[date(2014, 1, 31),
                         date(1999, 9, 24)],
                        [datetime(1998, 5, 26, 23, 33, 4),
                         datetime(2014, 2, 28, 13, 5, 13)]],
                       index=["DATE", "DATETIME"], columns=["X", "Y"])
        df_expected = DataFrame([[datetime(2014, 1, 31),
                                  datetime(1999, 9, 24)],
                                 [datetime(1998, 5, 26, 23, 33, 4),
                                  datetime(2014, 2, 28, 13, 5, 13)]],
                                index=["DATE", "DATETIME"], columns=["X", "Y"])

        with ensure_clean(ext) as filename2:
            writer1 = ExcelWriter(self.path)
            writer2 = ExcelWriter(filename2,
                                  date_format="DD.MM.YYYY",
                                  datetime_format="DD.MM.YYYY HH-MM-SS")

            df.to_excel(writer1, "test1")
            df.to_excel(writer2, "test1")

            writer1.close()
            writer2.close()

            reader1 = ExcelFile(self.path)
            reader2 = ExcelFile(filename2)

            rs1 = read_excel(reader1, "test1", index_col=0)
            rs2 = read_excel(reader2, "test1", index_col=0)

            tm.assert_frame_equal(rs1, rs2)

            # Since the reader returns a datetime object for dates,
            # we need to use df_expected to check the result.
            tm.assert_frame_equal(rs2, df_expected)

    def test_to_excel_interval_no_labels(self, *_):
        # see gh-19242
        #
        # Test writing Interval without labels.
        frame = DataFrame(np.random.randint(-10, 10, size=(20, 1)),
                          dtype=np.int64)
        expected = frame.copy()

        frame["new"] = pd.cut(frame[0], 10)
        expected["new"] = pd.cut(expected[0], 10).astype(str)

        frame.to_excel(self.path, "test1")
        reader = ExcelFile(self.path)

        recons = read_excel(reader, "test1", index_col=0)
        tm.assert_frame_equal(expected, recons)

    def test_to_excel_interval_labels(self, *_):
        # see gh-19242
        #
        # Test writing Interval with labels.
        frame = DataFrame(np.random.randint(-10, 10, size=(20, 1)),
                          dtype=np.int64)
        expected = frame.copy()
        intervals = pd.cut(frame[0], 10, labels=["A", "B", "C", "D", "E",
                                                 "F", "G", "H", "I", "J"])
        frame["new"] = intervals
        expected["new"] = pd.Series(list(intervals))

        frame.to_excel(self.path, "test1")
        reader = ExcelFile(self.path)

        recons = read_excel(reader, "test1", index_col=0)
        tm.assert_frame_equal(expected, recons)

    def test_to_excel_timedelta(self, *_):
        # see gh-19242, gh-9155
        #
        # Test writing timedelta to xls.
        frame = DataFrame(np.random.randint(-10, 10, size=(20, 1)),
                          columns=["A"], dtype=np.int64)
        expected = frame.copy()

        frame["new"] = frame["A"].apply(lambda x: timedelta(seconds=x))
        expected["new"] = expected["A"].apply(
            lambda x: timedelta(seconds=x).total_seconds() / float(86400))

        frame.to_excel(self.path, "test1")
        reader = ExcelFile(self.path)

        recons = read_excel(reader, "test1", index_col=0)
        tm.assert_frame_equal(expected, recons)

    def test_to_excel_periodindex(self, merge_cells, engine, ext):
        frame = self.tsframe
        xp = frame.resample('M', kind='period').mean()

        xp.to_excel(self.path, 'sht1')

        reader = ExcelFile(self.path)
        rs = read_excel(reader, 'sht1', index_col=0)
        tm.assert_frame_equal(xp, rs.to_period('M'))

    def test_to_excel_multiindex(self, merge_cells, engine, ext):
        frame = self.frame
        arrays = np.arange(len(frame.index) * 2).reshape(2, -1)
        new_index = MultiIndex.from_arrays(arrays,
                                           names=['first', 'second'])
        frame.index = new_index

        frame.to_excel(self.path, 'test1', header=False)
        frame.to_excel(self.path, 'test1', columns=['A', 'B'])

        # round trip
        frame.to_excel(self.path, 'test1', merge_cells=merge_cells)
        reader = ExcelFile(self.path)
        df = read_excel(reader, 'test1', index_col=[0, 1])
        tm.assert_frame_equal(frame, df)

    # GH13511
    def test_to_excel_multiindex_nan_label(self, merge_cells, engine, ext):
        frame = pd.DataFrame({'A': [None, 2, 3],
                              'B': [10, 20, 30],
                              'C': np.random.sample(3)})
        frame = frame.set_index(['A', 'B'])

        frame.to_excel(self.path, merge_cells=merge_cells)
        df = read_excel(self.path, index_col=[0, 1])
        tm.assert_frame_equal(frame, df)

    # Test for Issue 11328. If column indices are integers, make
    # sure they are handled correctly for either setting of
    # merge_cells
    def test_to_excel_multiindex_cols(self, merge_cells, engine, ext):
        frame = self.frame
        arrays = np.arange(len(frame.index) * 2).reshape(2, -1)
        new_index = MultiIndex.from_arrays(arrays,
                                           names=['first', 'second'])
        frame.index = new_index

        new_cols_index = MultiIndex.from_tuples([(40, 1), (40, 2),
                                                 (50, 1), (50, 2)])
        frame.columns = new_cols_index
        header = [0, 1]
        if not merge_cells:
            header = 0

        # round trip
        frame.to_excel(self.path, 'test1', merge_cells=merge_cells)
        reader = ExcelFile(self.path)
        df = read_excel(reader, 'test1', header=header,
                        index_col=[0, 1])
        if not merge_cells:
            fm = frame.columns.format(sparsify=False,
                                      adjoin=False, names=False)
            frame.columns = [".".join(map(str, q)) for q in zip(*fm)]
        tm.assert_frame_equal(frame, df)

    def test_to_excel_multiindex_dates(self, merge_cells, engine, ext):
        # try multiindex with dates
        tsframe = self.tsframe.copy()
        new_index = [tsframe.index, np.arange(len(tsframe.index))]
        tsframe.index = MultiIndex.from_arrays(new_index)

        tsframe.index.names = ['time', 'foo']
        tsframe.to_excel(self.path, 'test1', merge_cells=merge_cells)
        reader = ExcelFile(self.path)
        recons = read_excel(reader, 'test1',
                            index_col=[0, 1])

        tm.assert_frame_equal(tsframe, recons)
        assert recons.index.names == ('time', 'foo')

    def test_to_excel_multiindex_no_write_index(self, merge_cells, engine,
                                                ext):
        # Test writing and re-reading a MI witout the index. GH 5616.

        # Initial non-MI frame.
        frame1 = DataFrame({'a': [10, 20], 'b': [30, 40], 'c': [50, 60]})

        # Add a MI.
        frame2 = frame1.copy()
        multi_index = MultiIndex.from_tuples([(70, 80), (90, 100)])
        frame2.index = multi_index

        # Write out to Excel without the index.
        frame2.to_excel(self.path, 'test1', index=False)

        # Read it back in.
        reader = ExcelFile(self.path)
        frame3 = read_excel(reader, 'test1')

        # Test that it is the same as the initial frame.
        tm.assert_frame_equal(frame1, frame3)

    def test_to_excel_float_format(self, *_):
        df = DataFrame([[0.123456, 0.234567, 0.567567],
                        [12.32112, 123123.2, 321321.2]],
                       index=["A", "B"], columns=["X", "Y", "Z"])
        df.to_excel(self.path, "test1", float_format="%.2f")

        reader = ExcelFile(self.path)
        result = read_excel(reader, "test1", index_col=0)

        expected = DataFrame([[0.12, 0.23, 0.57],
                              [12.32, 123123.20, 321321.20]],
                             index=["A", "B"], columns=["X", "Y", "Z"])
        tm.assert_frame_equal(result, expected)

    def test_to_excel_output_encoding(self, merge_cells, engine, ext):
        # Avoid mixed inferred_type.
        df = DataFrame([[u"\u0192", u"\u0193", u"\u0194"],
                        [u"\u0195", u"\u0196", u"\u0197"]],
                       index=[u"A\u0192", u"B"],
                       columns=[u"X\u0193", u"Y", u"Z"])

        with ensure_clean("__tmp_to_excel_float_format__." + ext) as filename:
            df.to_excel(filename, sheet_name="TestSheet", encoding="utf8")
            result = read_excel(filename, "TestSheet",
                                encoding="utf8", index_col=0)
            tm.assert_frame_equal(result, df)

    def test_to_excel_unicode_filename(self, merge_cells, engine, ext):
        with ensure_clean(u("\u0192u.") + ext) as filename:
            try:
                f = open(filename, "wb")
            except UnicodeEncodeError:
                pytest.skip("No unicode file names on this system")
            else:
                f.close()

            df = DataFrame([[0.123456, 0.234567, 0.567567],
                            [12.32112, 123123.2, 321321.2]],
                           index=["A", "B"], columns=["X", "Y", "Z"])
            df.to_excel(filename, "test1", float_format="%.2f")

            reader = ExcelFile(filename)
            result = read_excel(reader, "test1", index_col=0)

            expected = DataFrame([[0.12, 0.23, 0.57],
                                  [12.32, 123123.20, 321321.20]],
                                 index=["A", "B"], columns=["X", "Y", "Z"])
            tm.assert_frame_equal(result, expected)

    # def test_to_excel_header_styling_xls(self, merge_cells, engine, ext):

    #     import StringIO
    #     s = StringIO(
    #     """Date,ticker,type,value
    #     2001-01-01,x,close,12.2
    #     2001-01-01,x,open ,12.1
    #     2001-01-01,y,close,12.2
    #     2001-01-01,y,open ,12.1
    #     2001-02-01,x,close,12.2
    #     2001-02-01,x,open ,12.1
    #     2001-02-01,y,close,12.2
    #     2001-02-01,y,open ,12.1
    #     2001-03-01,x,close,12.2
    #     2001-03-01,x,open ,12.1
    #     2001-03-01,y,close,12.2
    #     2001-03-01,y,open ,12.1""")
    #     df = read_csv(s, parse_dates=["Date"])
    #     pdf = df.pivot_table(values="value", rows=["ticker"],
    #                                          cols=["Date", "type"])

    #     try:
    #         import xlwt
    #         import xlrd
    #     except ImportError:
    #         pytest.skip

    #     filename = '__tmp_to_excel_header_styling_xls__.xls'
    #     pdf.to_excel(filename, 'test1')

    #     wbk = xlrd.open_workbook(filename,
    #                              formatting_info=True)
    #     assert ["test1"] == wbk.sheet_names()
    #     ws = wbk.sheet_by_name('test1')
    #     assert [(0, 1, 5, 7), (0, 1, 3, 5), (0, 1, 1, 3)] == ws.merged_cells
    #     for i in range(0, 2):
    #         for j in range(0, 7):
    #             xfx = ws.cell_xf_index(0, 0)
    #             cell_xf = wbk.xf_list[xfx]
    #             font = wbk.font_list
    #             assert 1 == font[cell_xf.font_index].bold
    #             assert 1 == cell_xf.border.top_line_style
    #             assert 1 == cell_xf.border.right_line_style
    #             assert 1 == cell_xf.border.bottom_line_style
    #             assert 1 == cell_xf.border.left_line_style
    #             assert 2 == cell_xf.alignment.hor_align
    #     os.remove(filename)
    # def test_to_excel_header_styling_xlsx(self, merge_cells, engine, ext):
    #     import StringIO
    #     s = StringIO(
    #     """Date,ticker,type,value
    #     2001-01-01,x,close,12.2
    #     2001-01-01,x,open ,12.1
    #     2001-01-01,y,close,12.2
    #     2001-01-01,y,open ,12.1
    #     2001-02-01,x,close,12.2
    #     2001-02-01,x,open ,12.1
    #     2001-02-01,y,close,12.2
    #     2001-02-01,y,open ,12.1
    #     2001-03-01,x,close,12.2
    #     2001-03-01,x,open ,12.1
    #     2001-03-01,y,close,12.2
    #     2001-03-01,y,open ,12.1""")
    #     df = read_csv(s, parse_dates=["Date"])
    #     pdf = df.pivot_table(values="value", rows=["ticker"],
    #                                          cols=["Date", "type"])
    #     try:
    #         import openpyxl
    #         from openpyxl.cell import get_column_letter
    #     except ImportError:
    #         pytest.skip
    #     if openpyxl.__version__ < '1.6.1':
    #         pytest.skip
    #     # test xlsx_styling
    #     filename = '__tmp_to_excel_header_styling_xlsx__.xlsx'
    #     pdf.to_excel(filename, 'test1')
    #     wbk = openpyxl.load_workbook(filename)
    #     assert ["test1"] == wbk.get_sheet_names()
    #     ws = wbk.get_sheet_by_name('test1')
    #     xlsaddrs = ["%s2" % chr(i) for i in range(ord('A'), ord('H'))]
    #     xlsaddrs += ["A%s" % i for i in range(1, 6)]
    #     xlsaddrs += ["B1", "D1", "F1"]
    #     for xlsaddr in xlsaddrs:
    #         cell = ws.cell(xlsaddr)
    #         assert cell.style.font.bold
    #         assert (openpyxl.style.Border.BORDER_THIN ==
    #                 cell.style.borders.top.border_style)
    #         assert (openpyxl.style.Border.BORDER_THIN ==
    #                 cell.style.borders.right.border_style)
    #         assert (openpyxl.style.Border.BORDER_THIN ==
    #                 cell.style.borders.bottom.border_style)
    #         assert (openpyxl.style.Border.BORDER_THIN ==
    #                 cell.style.borders.left.border_style)
    #         assert (openpyxl.style.Alignment.HORIZONTAL_CENTER ==
    #                 cell.style.alignment.horizontal)
    #     mergedcells_addrs = ["C1", "E1", "G1"]
    #     for maddr in mergedcells_addrs:
    #         assert ws.cell(maddr).merged
    #     os.remove(filename)

    @pytest.mark.parametrize("use_headers", [True, False])
    @pytest.mark.parametrize("r_idx_nlevels", [1, 2, 3])
    @pytest.mark.parametrize("c_idx_nlevels", [1, 2, 3])
    def test_excel_010_hemstring(self, merge_cells, engine, ext,
                                 c_idx_nlevels, r_idx_nlevels, use_headers):

        def roundtrip(data, header=True, parser_hdr=0, index=True):
            data.to_excel(self.path, header=header,
                          merge_cells=merge_cells, index=index)

            xf = ExcelFile(self.path)
            return read_excel(xf, xf.sheet_names[0], header=parser_hdr)

        # Basic test.
        parser_header = 0 if use_headers else None
        res = roundtrip(DataFrame([0]), use_headers, parser_header)

        assert res.shape == (1, 2)
        assert res.iloc[0, 0] is not np.nan

        # More complex tests with multi-index.
        nrows = 5
        ncols = 3

        # ensure limited functionality in 0.10
        # override of gh-2370 until sorted out in 0.11

        df = mkdf(nrows, ncols, r_idx_nlevels=r_idx_nlevels,
                  c_idx_nlevels=c_idx_nlevels)

        # This if will be removed once multi-column Excel writing
        # is implemented. For now fixing gh-9794.
        if c_idx_nlevels > 1:
            with pytest.raises(NotImplementedError):
                roundtrip(df, use_headers, index=False)
        else:
            res = roundtrip(df, use_headers)

            if use_headers:
                assert res.shape == (nrows, ncols + r_idx_nlevels)
            else:
                # First row taken as columns.
                assert res.shape == (nrows - 1, ncols + r_idx_nlevels)

            # No NaNs.
            for r in range(len(res.index)):
                for c in range(len(res.columns)):
                    assert res.iloc[r, c] is not np.nan

    def test_duplicated_columns(self, *_):
        # see gh-5235
        df = DataFrame([[1, 2, 3], [1, 2, 3], [1, 2, 3]],
                       columns=["A", "B", "B"])
        df.to_excel(self.path, "test1")
        expected = DataFrame([[1, 2, 3], [1, 2, 3], [1, 2, 3]],
                             columns=["A", "B", "B.1"])

        # By default, we mangle.
        result = read_excel(self.path, "test1", index_col=0)
        tm.assert_frame_equal(result, expected)

        # Explicitly, we pass in the parameter.
        result = read_excel(self.path, "test1", index_col=0,
                            mangle_dupe_cols=True)
        tm.assert_frame_equal(result, expected)

        # see gh-11007, gh-10970
        df = DataFrame([[1, 2, 3, 4], [5, 6, 7, 8]],
                       columns=["A", "B", "A", "B"])
        df.to_excel(self.path, "test1")

        result = read_excel(self.path, "test1", index_col=0)
        expected = DataFrame([[1, 2, 3, 4], [5, 6, 7, 8]],
                             columns=["A", "B", "A.1", "B.1"])
        tm.assert_frame_equal(result, expected)

        # see gh-10982
        df.to_excel(self.path, "test1", index=False, header=False)
        result = read_excel(self.path, "test1", header=None)

        expected = DataFrame([[1, 2, 3, 4], [5, 6, 7, 8]])
        tm.assert_frame_equal(result, expected)

        msg = "Setting mangle_dupe_cols=False is not supported yet"
        with pytest.raises(ValueError, match=msg):
            read_excel(self.path, "test1", header=None, mangle_dupe_cols=False)

    def test_swapped_columns(self, merge_cells, engine, ext):
        # Test for issue #5427.
        write_frame = DataFrame({'A': [1, 1, 1],
                                 'B': [2, 2, 2]})
        write_frame.to_excel(self.path, 'test1', columns=['B', 'A'])

        read_frame = read_excel(self.path, 'test1', header=0)

        tm.assert_series_equal(write_frame['A'], read_frame['A'])
        tm.assert_series_equal(write_frame['B'], read_frame['B'])

    def test_invalid_columns(self, *_):
        # see gh-10982
        write_frame = DataFrame({"A": [1, 1, 1],
                                 "B": [2, 2, 2]})

        with tm.assert_produces_warning(FutureWarning,
                                        check_stacklevel=False):
            write_frame.to_excel(self.path, "test1", columns=["B", "C"])

        expected = write_frame.reindex(columns=["B", "C"])
        read_frame = read_excel(self.path, "test1", index_col=0)
        tm.assert_frame_equal(expected, read_frame)

        with pytest.raises(KeyError):
            write_frame.to_excel(self.path, "test1", columns=["C", "D"])

    def test_comment_arg(self, *_):
        # see gh-18735
        #
        # Test the comment argument functionality to read_excel.

        # Create file to read in.
        df = DataFrame({"A": ["one", "#one", "one"],
                        "B": ["two", "two", "#two"]})
        df.to_excel(self.path, "test_c")

        # Read file without comment arg.
        result1 = read_excel(self.path, "test_c", index_col=0)

        result1.iloc[1, 0] = None
        result1.iloc[1, 1] = None
        result1.iloc[2, 1] = None

        result2 = read_excel(self.path, "test_c", comment="#", index_col=0)
        tm.assert_frame_equal(result1, result2)

    def test_comment_default(self, merge_cells, engine, ext):
        # Re issue #18735
        # Test the comment argument default to read_excel

        # Create file to read in
        df = DataFrame({'A': ['one', '#one', 'one'],
                        'B': ['two', 'two', '#two']})
        df.to_excel(self.path, 'test_c')

        # Read file with default and explicit comment=None
        result1 = read_excel(self.path, 'test_c')
        result2 = read_excel(self.path, 'test_c', comment=None)
        tm.assert_frame_equal(result1, result2)

    def test_comment_used(self, *_):
        # see gh-18735
        #
        # Test the comment argument is working as expected when used.

        # Create file to read in.
        df = DataFrame({"A": ["one", "#one", "one"],
                        "B": ["two", "two", "#two"]})
        df.to_excel(self.path, "test_c")

        # Test read_frame_comment against manually produced expected output.
        expected = DataFrame({"A": ["one", None, "one"],
                              "B": ["two", None, None]})
        result = read_excel(self.path, "test_c", comment="#", index_col=0)
        tm.assert_frame_equal(result, expected)

    def test_comment_empty_line(self, merge_cells, engine, ext):
        # Re issue #18735
        # Test that read_excel ignores commented lines at the end of file

        df = DataFrame({'a': ['1', '#2'], 'b': ['2', '3']})
        df.to_excel(self.path, index=False)

        # Test that all-comment lines at EoF are ignored
        expected = DataFrame({'a': [1], 'b': [2]})
        result = read_excel(self.path, comment='#')
        tm.assert_frame_equal(result, expected)

    def test_datetimes(self, merge_cells, engine, ext):

        # Test writing and reading datetimes. For issue #9139. (xref #9185)
        datetimes = [datetime(2013, 1, 13, 1, 2, 3),
                     datetime(2013, 1, 13, 2, 45, 56),
                     datetime(2013, 1, 13, 4, 29, 49),
                     datetime(2013, 1, 13, 6, 13, 42),
                     datetime(2013, 1, 13, 7, 57, 35),
                     datetime(2013, 1, 13, 9, 41, 28),
                     datetime(2013, 1, 13, 11, 25, 21),
                     datetime(2013, 1, 13, 13, 9, 14),
                     datetime(2013, 1, 13, 14, 53, 7),
                     datetime(2013, 1, 13, 16, 37, 0),
                     datetime(2013, 1, 13, 18, 20, 52)]

        write_frame = DataFrame({'A': datetimes})
        write_frame.to_excel(self.path, 'Sheet1')
        read_frame = read_excel(self.path, 'Sheet1', header=0)

        tm.assert_series_equal(write_frame['A'], read_frame['A'])

    def test_bytes_io(self, merge_cells, engine, ext):
        # see gh-7074
        bio = BytesIO()
        df = DataFrame(np.random.randn(10, 2))

        # Pass engine explicitly, as there is no file path to infer from.
        writer = ExcelWriter(bio, engine=engine)
        df.to_excel(writer)
        writer.save()

        bio.seek(0)
        reread_df = read_excel(bio, index_col=0)
        tm.assert_frame_equal(df, reread_df)

    def test_write_lists_dict(self, *_):
        # see gh-8188.
        df = DataFrame({"mixed": ["a", ["b", "c"], {"d": "e", "f": 2}],
                        "numeric": [1, 2, 3.0],
                        "str": ["apple", "banana", "cherry"]})
        df.to_excel(self.path, "Sheet1")
        read = read_excel(self.path, "Sheet1", header=0, index_col=0)

        expected = df.copy()
        expected.mixed = expected.mixed.apply(str)
        expected.numeric = expected.numeric.astype("int64")

        tm.assert_frame_equal(read, expected)

    def test_true_and_false_value_options(self, *_):
        # see gh-13347
        df = pd.DataFrame([["foo", "bar"]], columns=["col1", "col2"])
        expected = df.replace({"foo": True, "bar": False})

        df.to_excel(self.path)
        read_frame = read_excel(self.path, true_values=["foo"],
                                false_values=["bar"], index_col=0)
        tm.assert_frame_equal(read_frame, expected)

    def test_freeze_panes(self, *_):
        # see gh-15160
        expected = DataFrame([[1, 2], [3, 4]], columns=["col1", "col2"])
        expected.to_excel(self.path, "Sheet1", freeze_panes=(1, 1))

        result = read_excel(self.path, index_col=0)
        tm.assert_frame_equal(result, expected)

    def test_path_path_lib(self, merge_cells, engine, ext):
        df = tm.makeDataFrame()
        writer = partial(df.to_excel, engine=engine)

        reader = partial(pd.read_excel, index_col=0)
        result = tm.round_trip_pathlib(writer, reader,
                                       path="foo.{ext}".format(ext=ext))
        tm.assert_frame_equal(result, df)

    def test_path_local_path(self, merge_cells, engine, ext):
        df = tm.makeDataFrame()
        writer = partial(df.to_excel, engine=engine)

        reader = partial(pd.read_excel, index_col=0)
        result = tm.round_trip_pathlib(writer, reader,
                                       path="foo.{ext}".format(ext=ext))
        tm.assert_frame_equal(result, df)


@td.skip_if_no('openpyxl')
@pytest.mark.parametrize("merge_cells,ext,engine", [
    (None, '.xlsx', 'openpyxl')])
class TestOpenpyxlTests(_WriterBase):

    def test_to_excel_styleconverter(self, merge_cells, ext, engine):
        from openpyxl import styles

        hstyle = {
            "font": {
                "color": '00FF0000',
                "bold": True,
            },
            "borders": {
                "top": "thin",
                "right": "thin",
                "bottom": "thin",
                "left": "thin",
            },
            "alignment": {
                "horizontal": "center",
                "vertical": "top",
            },
            "fill": {
                "patternType": 'solid',
                'fgColor': {
                    'rgb': '006666FF',
                    'tint': 0.3,
                },
            },
            "number_format": {
                "format_code": "0.00"
            },
            "protection": {
                "locked": True,
                "hidden": False,
            },
        }

        font_color = styles.Color('00FF0000')
        font = styles.Font(bold=True, color=font_color)
        side = styles.Side(style=styles.borders.BORDER_THIN)
        border = styles.Border(top=side, right=side, bottom=side, left=side)
        alignment = styles.Alignment(horizontal='center', vertical='top')
        fill_color = styles.Color(rgb='006666FF', tint=0.3)
        fill = styles.PatternFill(patternType='solid', fgColor=fill_color)

        number_format = '0.00'

        protection = styles.Protection(locked=True, hidden=False)

        kw = _OpenpyxlWriter._convert_to_style_kwargs(hstyle)
        assert kw['font'] == font
        assert kw['border'] == border
        assert kw['alignment'] == alignment
        assert kw['fill'] == fill
        assert kw['number_format'] == number_format
        assert kw['protection'] == protection

    def test_write_cells_merge_styled(self, merge_cells, ext, engine):
        from pandas.io.formats.excel import ExcelCell

        sheet_name = 'merge_styled'

        sty_b1 = {'font': {'color': '00FF0000'}}
        sty_a2 = {'font': {'color': '0000FF00'}}

        initial_cells = [
            ExcelCell(col=1, row=0, val=42, style=sty_b1),
            ExcelCell(col=0, row=1, val=99, style=sty_a2),
        ]

        sty_merged = {'font': {'color': '000000FF', 'bold': True}}
        sty_kwargs = _OpenpyxlWriter._convert_to_style_kwargs(sty_merged)
        openpyxl_sty_merged = sty_kwargs['font']
        merge_cells = [
            ExcelCell(col=0, row=0, val='pandas',
                      mergestart=1, mergeend=1, style=sty_merged),
        ]

        with ensure_clean(ext) as path:
            writer = _OpenpyxlWriter(path)
            writer.write_cells(initial_cells, sheet_name=sheet_name)
            writer.write_cells(merge_cells, sheet_name=sheet_name)

            wks = writer.sheets[sheet_name]
            xcell_b1 = wks['B1']
            xcell_a2 = wks['A2']
            assert xcell_b1.font == openpyxl_sty_merged
            assert xcell_a2.font == openpyxl_sty_merged

    @pytest.mark.parametrize("mode,expected", [
        ('w', ['baz']), ('a', ['foo', 'bar', 'baz'])])
    def test_write_append_mode(self, merge_cells, ext, engine, mode, expected):
        import openpyxl
        df = DataFrame([1], columns=['baz'])

        with ensure_clean(ext) as f:
            wb = openpyxl.Workbook()
            wb.worksheets[0].title = 'foo'
            wb.worksheets[0]['A1'].value = 'foo'
            wb.create_sheet('bar')
            wb.worksheets[1]['A1'].value = 'bar'
            wb.save(f)

            writer = ExcelWriter(f, engine=engine, mode=mode)
            df.to_excel(writer, sheet_name='baz', index=False)
            writer.save()

            wb2 = openpyxl.load_workbook(f)
            result = [sheet.title for sheet in wb2.worksheets]
            assert result == expected

            for index, cell_value in enumerate(expected):
                assert wb2.worksheets[index]['A1'].value == cell_value


@td.skip_if_no('xlwt')
@pytest.mark.parametrize("merge_cells,ext,engine", [
    (None, '.xls', 'xlwt')])
class TestXlwtTests(_WriterBase):

    def test_excel_raise_error_on_multiindex_columns_and_no_index(
            self, merge_cells, ext, engine):
        # MultiIndex as columns is not yet implemented 9794
        cols = MultiIndex.from_tuples([('site', ''),
                                       ('2014', 'height'),
                                       ('2014', 'weight')])
        df = DataFrame(np.random.randn(10, 3), columns=cols)
        with pytest.raises(NotImplementedError):
            with ensure_clean(ext) as path:
                df.to_excel(path, index=False)

    def test_excel_multiindex_columns_and_index_true(self, merge_cells, ext,
                                                     engine):
        cols = MultiIndex.from_tuples([('site', ''),
                                       ('2014', 'height'),
                                       ('2014', 'weight')])
        df = pd.DataFrame(np.random.randn(10, 3), columns=cols)
        with ensure_clean(ext) as path:
            df.to_excel(path, index=True)

    def test_excel_multiindex_index(self, merge_cells, ext, engine):
        # MultiIndex as index works so assert no error #9794
        cols = MultiIndex.from_tuples([('site', ''),
                                       ('2014', 'height'),
                                       ('2014', 'weight')])
        df = DataFrame(np.random.randn(3, 10), index=cols)
        with ensure_clean(ext) as path:
            df.to_excel(path, index=False)

    def test_to_excel_styleconverter(self, merge_cells, ext, engine):
        import xlwt

        hstyle = {"font": {"bold": True},
                  "borders": {"top": "thin",
                              "right": "thin",
                              "bottom": "thin",
                              "left": "thin"},
                  "alignment": {"horizontal": "center", "vertical": "top"}}

        xls_style = _XlwtWriter._convert_to_style(hstyle)
        assert xls_style.font.bold
        assert xlwt.Borders.THIN == xls_style.borders.top
        assert xlwt.Borders.THIN == xls_style.borders.right
        assert xlwt.Borders.THIN == xls_style.borders.bottom
        assert xlwt.Borders.THIN == xls_style.borders.left
        assert xlwt.Alignment.HORZ_CENTER == xls_style.alignment.horz
        assert xlwt.Alignment.VERT_TOP == xls_style.alignment.vert

    def test_write_append_mode_raises(self, merge_cells, ext, engine):
        msg = "Append mode is not supported with xlwt!"

        with ensure_clean(ext) as f:
            with pytest.raises(ValueError, match=msg):
                ExcelWriter(f, engine=engine, mode='a')


@td.skip_if_no('xlsxwriter')
@pytest.mark.parametrize("merge_cells,ext,engine", [
    (None, '.xlsx', 'xlsxwriter')])
class TestXlsxWriterTests(_WriterBase):

    @td.skip_if_no('openpyxl')
    def test_column_format(self, merge_cells, ext, engine):
        # Test that column formats are applied to cells. Test for issue #9167.
        # Applicable to xlsxwriter only.
        with warnings.catch_warnings():
            # Ignore the openpyxl lxml warning.
            warnings.simplefilter("ignore")
            import openpyxl

        with ensure_clean(ext) as path:
            frame = DataFrame({'A': [123456, 123456],
                               'B': [123456, 123456]})

            writer = ExcelWriter(path)
            frame.to_excel(writer)

            # Add a number format to col B and ensure it is applied to cells.
            num_format = '#,##0'
            write_workbook = writer.book
            write_worksheet = write_workbook.worksheets()[0]
            col_format = write_workbook.add_format({'num_format': num_format})
            write_worksheet.set_column('B:B', None, col_format)
            writer.save()

            read_workbook = openpyxl.load_workbook(path)
            try:
                read_worksheet = read_workbook['Sheet1']
            except TypeError:
                # compat
                read_worksheet = read_workbook.get_sheet_by_name(name='Sheet1')

            # Get the number format from the cell.
            try:
                cell = read_worksheet['B2']
            except TypeError:
                # compat
                cell = read_worksheet.cell('B2')

            try:
                read_num_format = cell.number_format
            except Exception:
                read_num_format = cell.style.number_format._format_code

            assert read_num_format == num_format

    def test_write_append_mode_raises(self, merge_cells, ext, engine):
        msg = "Append mode is not supported with xlsxwriter!"

        with ensure_clean(ext) as f:
            with pytest.raises(ValueError, match=msg):
                ExcelWriter(f, engine=engine, mode='a')


class TestExcelWriterEngineTests(object):

    @pytest.mark.parametrize('klass,ext', [
        pytest.param(_XlsxWriter, '.xlsx', marks=pytest.mark.skipif(
            not td.safe_import('xlsxwriter'), reason='No xlsxwriter')),
        pytest.param(_OpenpyxlWriter, '.xlsx', marks=pytest.mark.skipif(
            not td.safe_import('openpyxl'), reason='No openpyxl')),
        pytest.param(_XlwtWriter, '.xls', marks=pytest.mark.skipif(
            not td.safe_import('xlwt'), reason='No xlwt'))
    ])
    def test_ExcelWriter_dispatch(self, klass, ext):
        with ensure_clean(ext) as path:
            writer = ExcelWriter(path)
            if ext == '.xlsx' and td.safe_import('xlsxwriter'):
                # xlsxwriter has preference over openpyxl if both installed
                assert isinstance(writer, _XlsxWriter)
            else:
                assert isinstance(writer, klass)

    def test_ExcelWriter_dispatch_raises(self):
        with pytest.raises(ValueError, match='No engine'):
            ExcelWriter('nothing')

    @pytest.mark.filterwarnings("ignore:\\nPanel:FutureWarning")
    def test_register_writer(self):
        # some awkward mocking to test out dispatch and such actually works
        called_save = []
        called_write_cells = []

        class DummyClass(ExcelWriter):
            called_save = False
            called_write_cells = False
            supported_extensions = ['test', 'xlsx', 'xls']
            engine = 'dummy'

            def save(self):
                called_save.append(True)

            def write_cells(self, *args, **kwargs):
                called_write_cells.append(True)

        def check_called(func):
            func()
            assert len(called_save) >= 1
            assert len(called_write_cells) >= 1
            del called_save[:]
            del called_write_cells[:]

        with pd.option_context('io.excel.xlsx.writer', 'dummy'):
            register_writer(DummyClass)
            writer = ExcelWriter('something.test')
            assert isinstance(writer, DummyClass)
            df = tm.makeCustomDataframe(1, 1)

            with catch_warnings(record=True):
                panel = tm.makePanel()
                func = lambda: df.to_excel('something.test')
                check_called(func)
                check_called(lambda: panel.to_excel('something.test'))
                check_called(lambda: df.to_excel('something.xlsx'))
                check_called(
                    lambda: df.to_excel(
                        'something.xls', engine='dummy'))


@pytest.mark.parametrize('engine', [
    pytest.param('xlwt',
                 marks=pytest.mark.xfail(reason='xlwt does not support '
                                                'openpyxl-compatible '
                                                'style dicts')),
    'xlsxwriter',
    'openpyxl',
])
def test_styler_to_excel(engine):
    def style(df):
        # XXX: RGB colors not supported in xlwt
        return DataFrame([['font-weight: bold', '', ''],
                          ['', 'color: blue', ''],
                          ['', '', 'text-decoration: underline'],
                          ['border-style: solid', '', ''],
                          ['', 'font-style: italic', ''],
                          ['', '', 'text-align: right'],
                          ['background-color: red', '', ''],
                          ['number-format: 0%', '', ''],
                          ['', '', ''],
                          ['', '', ''],
                          ['', '', '']],
                         index=df.index, columns=df.columns)

    def assert_equal_style(cell1, cell2):
        # XXX: should find a better way to check equality
        assert cell1.alignment.__dict__ == cell2.alignment.__dict__
        assert cell1.border.__dict__ == cell2.border.__dict__
        assert cell1.fill.__dict__ == cell2.fill.__dict__
        assert cell1.font.__dict__ == cell2.font.__dict__
        assert cell1.number_format == cell2.number_format
        assert cell1.protection.__dict__ == cell2.protection.__dict__

    def custom_converter(css):
        # use bold iff there is custom style attached to the cell
        if css.strip(' \n;'):
            return {'font': {'bold': True}}
        return {}

    pytest.importorskip('jinja2')
    pytest.importorskip(engine)

    # Prepare spreadsheets

    df = DataFrame(np.random.randn(11, 3))
    with ensure_clean('.xlsx' if engine != 'xlwt' else '.xls') as path:
        writer = ExcelWriter(path, engine=engine)
        df.to_excel(writer, sheet_name='frame')
        df.style.to_excel(writer, sheet_name='unstyled')
        styled = df.style.apply(style, axis=None)
        styled.to_excel(writer, sheet_name='styled')
        ExcelFormatter(styled, style_converter=custom_converter).write(
            writer, sheet_name='custom')
        writer.save()

        if engine not in ('openpyxl', 'xlsxwriter'):
            # For other engines, we only smoke test
            return
        openpyxl = pytest.importorskip('openpyxl')
        wb = openpyxl.load_workbook(path)

        # (1) compare DataFrame.to_excel and Styler.to_excel when unstyled
        n_cells = 0
        for col1, col2 in zip(wb['frame'].columns,
                              wb['unstyled'].columns):
            assert len(col1) == len(col2)
            for cell1, cell2 in zip(col1, col2):
                assert cell1.value == cell2.value
                assert_equal_style(cell1, cell2)
                n_cells += 1

        # ensure iteration actually happened:
        assert n_cells == (11 + 1) * (3 + 1)

        # (2) check styling with default converter

        # XXX: openpyxl (as at 2.4) prefixes colors with 00, xlsxwriter with FF
        alpha = '00' if engine == 'openpyxl' else 'FF'

        n_cells = 0
        for col1, col2 in zip(wb['frame'].columns,
                              wb['styled'].columns):
            assert len(col1) == len(col2)
            for cell1, cell2 in zip(col1, col2):
                ref = '%s%d' % (cell2.column, cell2.row)
                # XXX: this isn't as strong a test as ideal; we should
                #      confirm that differences are exclusive
                if ref == 'B2':
                    assert not cell1.font.bold
                    assert cell2.font.bold
                elif ref == 'C3':
                    assert cell1.font.color.rgb != cell2.font.color.rgb
                    assert cell2.font.color.rgb == alpha + '0000FF'
                elif ref == 'D4':
                    # This fails with engine=xlsxwriter due to
                    # https://bitbucket.org/openpyxl/openpyxl/issues/800
                    if engine == 'xlsxwriter' \
                       and (LooseVersion(openpyxl.__version__) <
                            LooseVersion('2.4.6')):
                        pass
                    else:
                        assert cell1.font.underline != cell2.font.underline
                        assert cell2.font.underline == 'single'
                elif ref == 'B5':
                    assert not cell1.border.left.style
                    assert (cell2.border.top.style ==
                            cell2.border.right.style ==
                            cell2.border.bottom.style ==
                            cell2.border.left.style ==
                            'medium')
                elif ref == 'C6':
                    assert not cell1.font.italic
                    assert cell2.font.italic
                elif ref == 'D7':
                    assert (cell1.alignment.horizontal !=
                            cell2.alignment.horizontal)
                    assert cell2.alignment.horizontal == 'right'
                elif ref == 'B8':
                    assert cell1.fill.fgColor.rgb != cell2.fill.fgColor.rgb
                    assert cell1.fill.patternType != cell2.fill.patternType
                    assert cell2.fill.fgColor.rgb == alpha + 'FF0000'
                    assert cell2.fill.patternType == 'solid'
                elif ref == 'B9':
                    assert cell1.number_format == 'General'
                    assert cell2.number_format == '0%'
                else:
                    assert_equal_style(cell1, cell2)

                assert cell1.value == cell2.value
                n_cells += 1

        assert n_cells == (11 + 1) * (3 + 1)

        # (3) check styling with custom converter
        n_cells = 0
        for col1, col2 in zip(wb['frame'].columns,
                              wb['custom'].columns):
            assert len(col1) == len(col2)
            for cell1, cell2 in zip(col1, col2):
                ref = '%s%d' % (cell2.column, cell2.row)
                if ref in ('B2', 'C3', 'D4', 'B5', 'C6', 'D7', 'B8', 'B9'):
                    assert not cell1.font.bold
                    assert cell2.font.bold
                else:
                    assert_equal_style(cell1, cell2)

                assert cell1.value == cell2.value
                n_cells += 1

        assert n_cells == (11 + 1) * (3 + 1)


@td.skip_if_no('openpyxl')
@pytest.mark.skipif(not PY36, reason='requires fspath')
class TestFSPath(object):

    def test_excelfile_fspath(self):
        with tm.ensure_clean('foo.xlsx') as path:
            df = DataFrame({"A": [1, 2]})
            df.to_excel(path)
            xl = ExcelFile(path)
            result = os.fspath(xl)
            assert result == path

    def test_excelwriter_fspath(self):
        with tm.ensure_clean('foo.xlsx') as path:
            writer = ExcelWriter(path)
            assert os.fspath(writer) == str(path)
